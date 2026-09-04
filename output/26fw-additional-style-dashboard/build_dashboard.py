from __future__ import annotations

import html
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = ROOT / "output" / "26fw-additional-style-dashboard"
SOURCE = ROOT / "추가 데이터" / "26FW 추가 구성 스타일.xlsx"


def esc(value: object) -> str:
    if value is None:
        return ""
    return html.escape(str(value)).replace("\n", "<br>")


def money(value: int) -> str:
    return f"{int(value):,}원"


def qty(value: int) -> str:
    return f"{int(value):,}장"


def short_money(value: int) -> str:
    return f"{value / 100000000:.1f}억"


def pct(value: int, total: int) -> str:
    return f"{value / total * 100:.1f}%" if total else "0.0%"


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    ws = wb.active

    merged: dict[tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        value = ws.cell(rng.min_row, rng.min_col).value
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                merged[(row, col)] = value

    def cell(row: int, col: int) -> object:
        return merged.get((row, col), ws.cell(row, col).value)

    rows: list[dict[str, object]] = []
    for row_index in range(3, ws.max_row + 1):
        style = cell(row_index, 5)
        if not style:
            continue
        style_code = str(style)
        match = re.search(r"WA\d{4}([A-Z]+)", style_code)
        rows.append(
            {
                "style": style_code,
                "season": cell(row_index, 4),
                "intent": cell(row_index, 6),
                "reference": cell(row_index, 7),
                "description": cell(row_index, 8),
                "name": cell(row_index, 9),
                "price": int(cell(row_index, 10) or 0),
                "color": cell(row_index, 11),
                "color_code": cell(row_index, 12),
                "quantity": int(cell(row_index, 13) or 0),
                "amount": int(cell(row_index, 14) or 0),
                "delivery": cell(row_index, 15),
                "category": match.group(1) if match else "기타",
            }
        )

    styles: dict[str, dict[str, object]] = {}
    for row in rows:
        style_code = str(row["style"])
        item = styles.setdefault(
            style_code,
            {
                "style": style_code,
                "name": row["name"],
                "season": row["season"],
                "intent": row["intent"],
                "description": row["description"],
                "price": row["price"],
                "delivery": row["delivery"],
                "category": row["category"],
                "quantity": 0,
                "amount": 0,
                "colors": 0,
                "color_names": [],
            },
        )
        item["quantity"] = int(item["quantity"]) + int(row["quantity"])
        item["amount"] = int(item["amount"]) + int(row["amount"])
        item["colors"] = int(item["colors"]) + 1
        if row["color"]:
            item["color_names"].append(str(row["color"]))

    style_list = sorted(styles.values(), key=lambda item: int(item["amount"]), reverse=True)

    category = defaultdict(lambda: {"quantity": 0, "amount": 0, "styles": set(), "colors": 0})
    season = defaultdict(lambda: {"quantity": 0, "amount": 0, "styles": set(), "colors": 0})
    for row in rows:
        cat = str(row["category"])
        seas = str(row["season"])
        category[cat]["quantity"] += int(row["quantity"])
        category[cat]["amount"] += int(row["amount"])
        category[cat]["styles"].add(str(row["style"]))
        category[cat]["colors"] += 1
        season[seas]["quantity"] += int(row["quantity"])
        season[seas]["amount"] += int(row["amount"])
        season[seas]["styles"].add(str(row["style"]))
        season[seas]["colors"] += 1

    cat_list = sorted(
        [
            {
                "category": key,
                "quantity": value["quantity"],
                "amount": value["amount"],
                "styles": len(value["styles"]),
                "colors": value["colors"],
            }
            for key, value in category.items()
        ],
        key=lambda item: int(item["amount"]),
        reverse=True,
    )
    season_list = sorted(
        [
            {
                "season": key,
                "quantity": value["quantity"],
                "amount": value["amount"],
                "styles": len(value["styles"]),
                "colors": value["colors"],
            }
            for key, value in season.items()
        ],
        key=lambda item: str(item["season"]),
    )

    top_total_styles = int(cell(1, 2) or 0)
    top_total_colors = int(cell(1, 3) or 0)
    total_qty = int(cell(1, 13) or sum(int(row["quantity"]) for row in rows))
    total_amount = int(cell(1, 14) or sum(int(row["amount"]) for row in rows))
    unique_style_count = len(styles)
    row_color_count = len(rows)
    missing_color_name = sum(1 for row in rows if not row["color"])
    value_errors = sum(
        1
        for row in rows
        for key in ["reference", "description"]
        if str(row.get(key)) == "#VALUE!"
    )
    deliveries = sorted({str(row["delivery"]) for row in rows if row["delivery"]})

    def group_summary(codes: set[str]) -> dict[str, int]:
        selected = [styles[code] for code in codes if code in styles]
        return {
            "styles": len(selected),
            "qty": sum(int(item["quantity"]) for item in selected),
            "amount": sum(int(item["amount"]) for item in selected),
            "colors": sum(int(item["colors"]) for item in selected),
        }

    groups = {
        "리오더/팩 티셔츠": group_summary({"WA2603STE1", "WA2603STE3"}),
        "컬리지 그래픽 스웻": group_summary({"WA2603HZ15", "WA2603CR06", "WA2603HD06", "WA2603HZ06"}),
        "기모 코어 구성": group_summary({"WA2604HZ01", "WA2604HD02", "WA2604CR01", "WA2604PT01"}),
        "우먼스 보완": group_summary({"WA2603CR67", "WA2604CR68", "WA2604HZ68"}),
    }

    group_cards = ""
    group_colors = ["border-l-bcave-dark", "border-l-bcave-amber", "border-l-bcave-green", "border-l-bcave-medium"]
    for index, (name, group) in enumerate(groups.items(), start=1):
        group_cards += f"""
          <div class="card border-l-4 {group_colors[index - 1]}">
            <div class="text-xs font-black uppercase text-bcave-medium">Group {index}</div>
            <h3 class="mt-2 text-lg font-black text-bcave-dark">{esc(name)}</h3>
            <div class="mt-4 grid grid-cols-3 gap-2 text-center">
              <div class="rounded bg-bcave-bg p-2"><div class="text-xs font-bold text-gray-400">STYLE</div><div class="font-black text-bcave-dark">{group["styles"]}</div></div>
              <div class="rounded bg-bcave-bg p-2"><div class="text-xs font-bold text-gray-400">QTY</div><div class="font-black text-bcave-dark">{group["qty"]:,}</div></div>
              <div class="rounded bg-bcave-bg p-2"><div class="text-xs font-bold text-gray-400">AMT</div><div class="font-black text-bcave-dark">{short_money(group["amount"])}</div></div>
            </div>
            <p class="mt-3 text-xs font-semibold leading-5 text-gray-500">전체 금액 비중 {pct(group["amount"], total_amount)}, 컬러 {group["colors"]}개 기준</p>
          </div>
        """

    image_refs = [
        {
            "target": "WA2603STE1",
            "title": "2PACK 시그니처 반팔 티셔츠",
            "note": "26FW 2PACK 리오더/상시 아이템 참고",
            "img": "assets/ref_2pack_signature.jpg",
            "source": "Wacky Willy 공식몰",
            "url": "https://wackywilly.co.kr/product/detail.html?product_no=16036",
        },
        {
            "target": "WA2603STE3",
            "title": "3PACK 시그니처 반팔 티셔츠",
            "note": "26FW 3PACK 리오더/상시 아이템 참고",
            "img": "assets/ref_3pack_signature.jpg",
            "source": "Wacky Willy 공식몰",
            "url": "https://wackywilly.co.kr/product/26ss-3pack-%EC%8B%9C%EA%B7%B8%EB%8B%88%EC%B2%98-%EB%B0%98%ED%8C%94-%ED%8B%B0%EC%85%94%EC%B8%A0-%EB%B8%94%EB%9E%99%ED%99%94%EC%9D%B4%ED%8A%B8%EA%B7%B8%EB%A0%88%EC%9D%B4/12905/",
        },
        {
            "target": "WA2604PT01",
            "title": "기모 스웻 팬츠",
            "note": "WA2504PT01 캐리오버 기반, 핏/디테일 변경 예정",
            "img": "assets/ref_wa2504pt01.jpg",
            "source": "Wacky Willy 공식몰",
            "url": "https://wackywilly.co.kr/product/detail.html?product_no=13048",
        },
        {
            "target": "WA2604CR01",
            "title": "기모 키키로고 맨투맨",
            "note": "WA2504CR01 캐리오버 기반, 로고/핏 변경 예정",
            "img": "assets/ref_wa2504cr01.jpg",
            "source": "Wacky Willy 공식몰 / 무신사",
            "url": "https://www.musinsa.com/products/5367005",
        },
        {
            "target": "WA2604HD02",
            "title": "기모 타이포 그래픽 후드",
            "note": "WA2504HD11 캐리오버 기반, 원단 컬러 변경",
            "img": "assets/ref_wa2504hd11.jpg",
            "source": "Wacky Willy 공식몰",
            "url": "https://wackywilly.co.kr/product/detail.html?product_no=12962",
        },
        {
            "target": "WA2603HZ15",
            "title": "스케치 그래픽 후드 스웻",
            "note": "WA2503HD11 그래픽을 후드집업으로 전환하는 참고 아트웍",
            "img": "assets/ref_wa2503hd11.jpg",
            "source": "Wacky Willy 공식몰",
            "url": "https://wackywilly.co.kr/product/detail.html?product_no=12428",
        },
    ]
    image_cards = ""
    for image in image_refs:
        image_cards += f"""
          <article class="overflow-hidden rounded-lg border border-bcave-light bg-white">
            <div class="bg-bcave-bg p-3">
              <img src="{image["img"]}" alt="{esc(image["title"])}" class="product-img" />
            </div>
            <div class="p-4">
              <div class="text-xs font-black text-bcave-medium">REFERENCE FOR {esc(image["target"])}</div>
              <h3 class="mt-1 text-base font-black text-bcave-dark">{esc(image["title"])}</h3>
              <p class="mt-2 text-xs font-semibold leading-5 text-gray-500">{esc(image["note"])}</p>
              <a href="{image["url"]}" target="_blank" rel="noreferrer" class="mt-3 inline-flex text-xs font-black text-bcave-red">{esc(image["source"])}</a>
            </div>
          </article>
        """

    style_rows = ""
    for index, style in enumerate(style_list, start=1):
        badge = "P1" if index <= 4 else ("P2" if index <= 9 else "P3")
        badge_class = "bg-bcave-dark text-white" if badge == "P1" else ("bg-bcave-amber/15 text-bcave-dark" if badge == "P2" else "bg-bcave-light text-bcave-medium")
        colors = ", ".join(style["color_names"]) if style["color_names"] else "컬러명 확인 필요"
        color_class = "text-bcave-red font-extrabold" if not style["color_names"] else "text-gray-600"
        style_rows += f"""
          <tr class="border-b border-bcave-light/80 align-top hover:bg-bcave-bg">
            <td class="px-4 py-3"><span class="rounded px-2 py-1 text-xs font-black {badge_class}">{badge}</span></td>
            <td class="px-4 py-3 font-black text-bcave-dark">{esc(style["style"])}<div class="mt-1 text-xs font-semibold text-bcave-medium">{esc(style["category"])} · {esc(style["season"])} · {esc(style["delivery"])}</div></td>
            <td class="px-4 py-3 font-semibold text-gray-700">{esc(style["name"])}</td>
            <td class="px-4 py-3 text-right font-black text-bcave-dark">{qty(int(style["quantity"]))}<div class="mt-1 text-xs font-semibold text-gray-400">{style["colors"]} colors</div></td>
            <td class="px-4 py-3 text-right font-black text-bcave-dark">{money(int(style["amount"]))}<div class="mt-1 text-xs font-semibold text-gray-400">{pct(int(style["amount"]), total_amount)}</div></td>
            <td class="px-4 py-3 text-sm {color_class}">{esc(colors)}</td>
          </tr>
        """

    cat_rows = ""
    for cat in cat_list:
        cat_rows += f"""
          <tr class="border-b border-bcave-light/80">
            <td class="px-4 py-3 font-black text-bcave-dark">{esc(cat["category"])}</td>
            <td class="px-4 py-3 text-right font-bold">{cat["styles"]}개</td>
            <td class="px-4 py-3 text-right font-bold">{cat["colors"]}개</td>
            <td class="px-4 py-3 text-right font-black text-bcave-dark">{qty(int(cat["quantity"]))}</td>
            <td class="px-4 py-3 text-right font-black text-bcave-dark">{money(int(cat["amount"]))}</td>
            <td class="px-4 py-3 text-right font-bold text-bcave-medium">{pct(int(cat["amount"]), total_amount)}</td>
          </tr>
        """

    season_cards = "".join(
        f"""
          <div class="rounded bg-bcave-bg p-3">
            <div class="font-black text-bcave-dark">{esc(item["season"])}</div>
            <div class="text-xs font-bold text-gray-500">{int(item["quantity"]):,}장 · {short_money(int(item["amount"]))}</div>
          </div>
        """
        for item in season_list
    )

    cat_labels = json.dumps([item["category"] for item in cat_list], ensure_ascii=False)
    cat_amounts = json.dumps([round(int(item["amount"]) / 100000000, 2) for item in cat_list])
    cat_qtys = json.dumps([int(item["quantity"]) for item in cat_list])
    style_labels = json.dumps([item["style"] for item in style_list], ensure_ascii=False)
    style_amounts = json.dumps([round(int(item["amount"]) / 100000000, 2) for item in style_list])
    style_qtys = json.dumps([int(item["quantity"]) for item in style_list])
    season_labels = json.dumps([str(item["season"]) for item in season_list], ensure_ascii=False)
    season_amounts = json.dumps([round(int(item["amount"]) / 100000000, 2) for item in season_list])
    season_qtys = json.dumps([int(item["quantity"]) for item in season_list])

    html_text = f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>B.CAVE - 26FW 추가 구성 스타일 대시보드</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  <link rel="preconnect" href="https://cdn.jsdelivr.net" />
  <link rel="stylesheet" crossorigin href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.8/dist/web/static/pretendard.css" />
  <script>
    tailwind.config = {{
      theme: {{
        extend: {{
          colors: {{
            bcave: {{ dark: '#264148', medium: '#718790', light: '#DEE2E3', bg: '#F8F9FA', red: '#B94735', amber: '#D4A64F', green: '#3D7465' }}
          }},
          fontFamily: {{ sans: ['Pretendard', 'sans-serif'] }},
          boxShadow: {{ report: '0 16px 44px rgba(38, 65, 72, 0.10)' }}
        }}
      }}
    }};
  </script>
  <style>
    :root {{ --bcave-dark:#264148; --bcave-medium:#718790; --bcave-light:#DEE2E3; --bcave-bg:#F8F9FA; }}
    * {{ box-sizing: border-box; }}
    body {{
      font-family: Pretendard, "Malgun Gothic", "Apple SD Gothic Neo", sans-serif;
      letter-spacing: 0;
      background: linear-gradient(135deg, rgba(248,249,250,.98), rgba(222,226,227,.80)), radial-gradient(circle at 8% 8%, rgba(38,65,72,.12), transparent 34%), radial-gradient(circle at 92% 12%, rgba(212,166,79,.13), transparent 30%);
      color: #1f2933;
    }}
    .card {{ background: rgba(255,255,255,.95); border: 1px solid rgba(38,65,72,.13); border-radius: 8px; box-shadow: 0 10px 30px rgba(38,65,72,.07); padding: 1.5rem; }}
    .pill {{ display:inline-flex; align-items:center; min-height:28px; padding:0 12px; border-radius:999px; font-size:12px; font-weight:900; }}
    .section-title {{ border-left: 5px solid #718790; padding-left: .875rem; color: #264148; font-weight: 900; }}
    .chart-box {{ position: relative; height: 20rem; }}
    .product-img {{ width: 100%; height: 210px; object-fit: contain; background: #F8F9FA; border-radius: 6px; display: block; }}
    @media print {{ body {{ background:#fff; }} .card {{ box-shadow:none; break-inside: avoid; }} .print-break {{ break-before: page; }} }}
  </style>
</head>
<body class="p-4 md:p-10">
  <div class="mx-auto max-w-7xl space-y-6">
    <header class="rounded-lg border border-bcave-dark/20 bg-white/90 p-6 shadow-report">
      <div class="flex flex-col gap-5 md:flex-row md:items-center md:justify-between">
        <div>
          <span class="pill bg-bcave-light text-bcave-dark">26FW 추가 구성 · 스타일 대시보드</span>
          <h1 class="mt-4 text-3xl font-black leading-tight text-bcave-dark md:text-4xl">26FW 추가 구성 스타일 검토 보고</h1>
          <p class="mt-2 text-base font-semibold text-bcave-medium">추가 구성 스타일의 수량·금액·품목군 집중도와 실행 리스크를 한 화면에서 점검합니다.</p>
        </div>
        <div class="flex items-center gap-5">
          <div class="hidden h-16 w-px bg-bcave-light md:block"></div>
          <img src="assets/bcave_logo.png" alt="B.CAVE Logo" class="h-10 w-auto" />
        </div>
      </div>
      <div class="mt-6 grid gap-3 border-t border-bcave-light pt-4 text-sm font-bold text-bcave-medium md:grid-cols-4">
        <div>Source <span class="block pt-1 text-bcave-dark">{esc(SOURCE.relative_to(ROOT).as_posix())}</span></div>
        <div>Sheet <span class="block pt-1 text-bcave-dark">{esc(ws.title)}</span></div>
        <div>Target Delivery <span class="block pt-1 text-bcave-dark">{esc(", ".join(deliveries))}</span></div>
        <div>Generated <span class="block pt-1 text-bcave-dark">2026-07-15</span></div>
      </div>
    </header>

    <section class="grid grid-cols-1 gap-4 md:grid-cols-5">
      <div class="card border-l-4 border-l-bcave-dark"><p class="text-xs font-black uppercase text-gray-500">상단 표기 스타일</p><p class="mt-2 text-2xl font-black text-bcave-dark">{top_total_styles}개</p><p class="mt-1 text-sm font-semibold text-gray-400">세부 확정품번 {unique_style_count}개</p></div>
      <div class="card border-l-4 border-l-bcave-medium"><p class="text-xs font-black uppercase text-gray-500">컬러 수</p><p class="mt-2 text-2xl font-black text-bcave-dark">{top_total_colors}개</p><p class="mt-1 text-sm font-semibold text-gray-400">세부 행 {row_color_count}개</p></div>
      <div class="card border-l-4 border-l-bcave-green"><p class="text-xs font-black uppercase text-gray-500">총 수량</p><p class="mt-2 text-2xl font-black text-bcave-dark">{qty(total_qty)}</p><p class="mt-1 text-sm font-semibold text-gray-400">컬러별 수량 합계</p></div>
      <div class="card border-l-4 border-l-bcave-amber"><p class="text-xs font-black uppercase text-gray-500">총 금액</p><p class="mt-2 text-2xl font-black text-bcave-dark">{short_money(total_amount)}</p><p class="mt-1 text-sm font-semibold text-gray-400">{money(total_amount)}</p></div>
      <div class="card border-l-4 border-l-bcave-red bg-red-50"><p class="text-xs font-black uppercase text-bcave-red">체크 필요</p><p class="mt-2 text-2xl font-black text-bcave-red">{missing_color_name + value_errors}건</p><p class="mt-1 text-sm font-bold text-red-500">색상명 누락 {missing_color_name} · 수식 오류 {value_errors}</p></div>
    </section>

    <section class="grid grid-cols-1 gap-6 lg:grid-cols-3">
      <div class="card lg:col-span-2"><div class="mb-4 flex flex-col gap-2 md:flex-row md:items-center md:justify-between"><h2 class="text-lg font-black text-bcave-dark">품목군별 금액·수량 집중도</h2><span class="pill bg-bcave-light text-bcave-medium">HZ/CR/STE 중심 구성</span></div><div class="chart-box"><canvas id="categoryChart"></canvas></div></div>
      <aside class="card bg-bcave-dark text-white">
        <h2 class="border-b border-white/20 pb-3 text-xl font-black text-bcave-light">Executive Summary</h2>
        <ul class="mt-5 space-y-4 text-sm leading-6">
          <li><strong class="block text-base text-white">1. 총 규모는 49,400장 / 40.7억</strong><span class="text-white/70">26FW 추가 구성은 수량 기준 STE와 CR/HZ가 볼륨을 만들고, 금액 기준 HZ·CR·STE가 핵심 축입니다.</span></li>
          <li><strong class="block text-base text-white">2. 9월 말 납기 일괄 관리 필요</strong><span class="text-white/70">모든 세부 행의 목표납기가 9월 말로 입력되어 있어 발주·원단·작업지시서 병목 관리가 중요합니다.</span></li>
          <li><strong class="block text-base text-white">3. 데이터 정합성 체크 필요</strong><span class="text-white/70">상단 총계는 11스타일이나 확정품번 기준은 13개입니다. 우먼스 일부 컬러명과 레퍼런스 수식 오류도 확인이 필요합니다.</span></li>
        </ul>
      </aside>
    </section>

    <section class="grid grid-cols-1 gap-6 lg:grid-cols-3">
      <div class="card lg:col-span-2"><div class="mb-4 flex flex-col gap-2 md:flex-row md:items-center md:justify-between"><h2 class="text-lg font-black text-bcave-dark">스타일별 금액 Top</h2><span class="pill bg-bcave-light text-bcave-medium">Tooltip: 수량·금액 동시 표기</span></div><div class="chart-box"><canvas id="styleChart"></canvas></div></div>
      <div class="card"><h2 class="text-lg font-black text-bcave-dark">시즌 구분 비중</h2><p class="mt-1 text-sm font-semibold text-bcave-medium">엑셀의 시즌 열 F/W 기준</p><div class="mt-4 h-64"><canvas id="seasonChart"></canvas></div><div class="mt-4 grid grid-cols-2 gap-3 text-center text-sm">{season_cards}</div></div>
    </section>

    <section class="space-y-4">
      <div class="flex flex-col gap-2 md:flex-row md:items-end md:justify-between">
        <h2 class="section-title text-2xl">상품 이미지 레퍼런스</h2>
        <p class="text-sm font-semibold text-bcave-medium">공개몰에서 확인 가능한 리오더/캐리오버/아트웍 참고 이미지입니다. 26FW 신규 확정 컷은 별도 확인이 필요합니다.</p>
      </div>
      <div class="grid grid-cols-1 gap-4 md:grid-cols-2 lg:grid-cols-3">{image_cards}</div>
    </section>

    <section class="space-y-4"><h2 class="section-title text-2xl">구성 그룹별 해석</h2><div class="grid grid-cols-1 gap-4 md:grid-cols-4">{group_cards}</div></section>

    <section class="grid grid-cols-1 gap-6 lg:grid-cols-2">
      <div class="card"><h2 class="text-lg font-black text-bcave-dark">품목군 요약</h2><div class="mt-4 overflow-x-auto"><table class="w-full min-w-[640px] text-sm"><thead class="bg-bcave-light text-xs font-black uppercase text-bcave-dark"><tr><th class="px-4 py-3 text-left">품목군</th><th class="px-4 py-3 text-right">스타일</th><th class="px-4 py-3 text-right">컬러</th><th class="px-4 py-3 text-right">수량</th><th class="px-4 py-3 text-right">금액</th><th class="px-4 py-3 text-right">비중</th></tr></thead><tbody>{cat_rows}</tbody></table></div></div>
      <div class="card border-2 border-bcave-dark"><h2 class="text-lg font-black text-bcave-dark">실행 우선순위</h2><div class="mt-4 space-y-3"><div class="rounded bg-bcave-bg p-4"><div class="text-xs font-black text-bcave-medium">P1 · 납기/생산</div><p class="mt-1 text-sm font-semibold leading-6 text-gray-700">전 스타일 목표납기가 9월 말입니다. 발주 확정, 원단 수급, 작업지시서, 생산 가능 수량을 한 번에 묶어 관리해야 합니다.</p></div><div class="rounded bg-bcave-bg p-4"><div class="text-xs font-black text-bcave-medium">P2 · 볼륨 검증</div><p class="mt-1 text-sm font-semibold leading-6 text-gray-700">STE 2PACK/3PACK은 16,000장으로 전체 수량의 {pct(groups["리오더/팩 티셔츠"]["qty"], total_qty)}입니다. 리오더/상시 수요 성격으로 별도 판매 속도 기준이 필요합니다.</p></div><div class="rounded bg-red-50 p-4"><div class="text-xs font-black text-bcave-red">P3 · 데이터 정합성</div><p class="mt-1 text-sm font-bold leading-6 text-gray-700">상단 11스타일과 확정품번 13개 차이, 우먼스 일부 컬러명 누락, #VALUE! 레퍼런스는 공유 전 원본 확인이 필요합니다.</p></div></div></div>
    </section>

    <section class="card print-break"><div class="mb-4 flex flex-col gap-2 md:flex-row md:items-center md:justify-between"><h2 class="text-xl font-black text-bcave-dark">스타일별 상세 리스트</h2><span class="pill bg-bcave-light text-bcave-medium">확정품번 기준 {unique_style_count}개 · 금액순</span></div><div class="overflow-x-auto"><table class="w-full min-w-[1080px] text-sm"><thead class="bg-bcave-light text-xs font-black uppercase text-bcave-dark"><tr><th class="px-4 py-3 text-left">Priority</th><th class="px-4 py-3 text-left">Style</th><th class="px-4 py-3 text-left">상품명</th><th class="px-4 py-3 text-right">수량</th><th class="px-4 py-3 text-right">금액</th><th class="px-4 py-3 text-left">컬러</th></tr></thead><tbody>{style_rows}</tbody></table></div></section>

    <section class="grid grid-cols-1 gap-6 lg:grid-cols-3">
      <div class="card bg-bcave-dark text-white lg:col-span-2"><h2 class="mb-4 text-xl font-black text-bcave-light">최종 보고 문안</h2><p class="text-sm font-medium leading-7 text-white/80 md:text-base">26FW 추가 구성 스타일은 엑셀 상단 총계 기준 <strong class="text-bcave-amber">11스타일 / 33컬러 / 49,400장 / 4,066,600,000원</strong> 규모로 확인됩니다. 세부 확정품번 기준으로는 13개 스타일이 집계되어 상단 총계와 차이가 있으므로, 공유 전 스타일 카운트 기준 확인이 필요합니다. 구성은 HZ·CR·STE 중심이며, 리오더 성격의 2PACK/3PACK과 컬리지 그래픽 스웻, 기모 코어 아이템이 물량과 금액의 핵심입니다. 모든 세부 행의 목표납기가 9월 말로 입력되어 있어 발주 확정, 원단 수급, 작업지시서 반영, 생산 가능 수량을 우선 관리해야 합니다.</p></div>
      <div class="card"><h2 class="text-lg font-black text-bcave-dark">Data Basis</h2><ul class="mt-4 list-disc space-y-2 pl-5 text-sm font-semibold leading-6 text-gray-600"><li>원천 파일: {esc(SOURCE.relative_to(ROOT).as_posix())}</li><li>시트: {esc(ws.title)}</li><li>집계 기준: 확정품번, 컬러별 수량, 컬러별 금액</li><li>제품 이미지는 원본에 검증 가능한 고해상도 파일이 없어 사용하지 않았습니다.</li><li>Teams 맥락: 26FW 추가 스타일 업데이트 및 9월 입고 목표 관리 항목과 연결해 해석했습니다.</li></ul></div>
    </section>
  </div>

  <script>
    const bcave = {{ dark: '#264148', medium: '#718790', light: '#DEE2E3', red: '#B94735', amber: '#D4A64F', green: '#3D7465' }};
    const fmtQty = (value) => new Intl.NumberFormat('ko-KR').format(value) + '장';
    const fmtAmt = (value) => value.toFixed(1) + '억';
    new Chart(document.getElementById('categoryChart'), {{ type: 'bar', data: {{ labels: {cat_labels}, datasets: [{{ label: '금액(억원)', data: {cat_amounts}, backgroundColor: [bcave.dark, bcave.medium, bcave.amber, bcave.green, bcave.light], borderRadius: 6, yAxisID: 'y' }}, {{ label: '수량(장)', data: {cat_qtys}, type: 'line', borderColor: bcave.red, backgroundColor: bcave.red, tension: .35, pointRadius: 4, yAxisID: 'y1' }}] }}, options: {{ responsive: true, maintainAspectRatio: false, plugins: {{ tooltip: {{ backgroundColor: bcave.dark, padding: 12 }} }}, scales: {{ y: {{ beginAtZero: true, title: {{ display: true, text: '금액(억원)' }}, grid: {{ color: '#E5E7EB' }} }}, y1: {{ beginAtZero: true, position: 'right', title: {{ display: true, text: '수량(장)' }}, grid: {{ drawOnChartArea: false }} }}, x: {{ grid: {{ display: false }} }} }} }} }});
    new Chart(document.getElementById('styleChart'), {{ type: 'bar', data: {{ labels: {style_labels}, datasets: [{{ label: '금액(억원)', data: {style_amounts}, backgroundColor: '#264148', borderRadius: 6, barThickness: 18 }}] }}, options: {{ indexAxis: 'y', responsive: true, maintainAspectRatio: false, plugins: {{ legend: {{ display: false }}, tooltip: {{ backgroundColor: bcave.dark, padding: 12, callbacks: {{ afterLabel(context) {{ const quantities = {style_qtys}; return '수량: ' + fmtQty(quantities[context.dataIndex]); }} }} }} }}, scales: {{ x: {{ beginAtZero: true, grid: {{ color: '#E5E7EB' }} }}, y: {{ grid: {{ display: false }}, ticks: {{ font: {{ weight: '800' }}, color: bcave.dark }} }} }} }} }});
    new Chart(document.getElementById('seasonChart'), {{ type: 'doughnut', data: {{ labels: {season_labels}, datasets: [{{ data: {season_amounts}, backgroundColor: [bcave.dark, bcave.amber], borderWidth: 0 }}] }}, options: {{ responsive: true, maintainAspectRatio: false, cutout: '62%', plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ family: 'Pretendard', weight: '800' }} }} }}, tooltip: {{ backgroundColor: bcave.dark, padding: 12, callbacks: {{ label(context) {{ const quantities = {season_qtys}; return context.label + ': ' + fmtAmt(context.raw) + ' / ' + fmtQty(quantities[context.dataIndex]); }} }} }} }} }} }});
  </script>
</body>
</html>
"""

    (OUT_DIR / "index.html").write_text(html_text, encoding="utf-8")
    print(OUT_DIR / "index.html")


if __name__ == "__main__":
    main()
