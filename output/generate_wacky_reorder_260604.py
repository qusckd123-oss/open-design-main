from __future__ import annotations

import datetime as dt
import re
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
from matplotlib import font_manager
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import load_workbook


TARGETS = [
    "ST32 PI",
    "ST27 EC",
    "ST16 WH",
    "ST16 BK",
    "ST34 GN",
    "ST34 WH",
    "ST31 YE",
    "ST31 CH",
    "ST31 WH",
    "ST24 NA",
    "ST24 CM",
]


@dataclass
class ReorderEvent:
    label: str
    qty: int
    date: dt.date | None


@dataclass
class StyleData:
    sheet: str
    sku: str
    review_date: dt.date
    inbound: int
    sales: int
    inventory: int
    reorder_request: int
    reorders: list[ReorderEvent]
    weeks: list[dt.date]
    labels: list[str]
    weekly_sales: list[float]
    cum_sales: list[float]
    rates_pct: list[float]


def set_font() -> None:
    preferred = ["Malgun Gothic", "Noto Sans CJK KR", "Noto Sans KR", "Apple SD Gothic Neo"]
    available = {font.name for font in font_manager.fontManager.ttflist}
    for name in preferred:
        if name in available:
            plt.rcParams["font.family"] = name
            break
    plt.rcParams["axes.unicode_minus"] = False


def as_int(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    match = re.search(r"-?\d[\d,]*(?:\.\d+)?", str(value))
    return int(round(float(match.group(0).replace(",", "")))) if match else 0


def as_float(value) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("%", "").strip()
    try:
        return float(text)
    except ValueError:
        return float(as_int(text))


def as_date(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if not value:
        return None
    text = str(value)
    match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if match:
        return dt.date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    match = re.search(r"(?<!\d)(\d{1,2})[-/.](\d{1,2})(?!\d)", text)
    if match:
        return dt.date(2026, int(match.group(1)), int(match.group(2)))
    return None


def col_name(index: int) -> str:
    out = ""
    while index:
        index, rem = divmod(index - 1, 26)
        out = chr(65 + rem) + out
    return out


def parse_reorders(ws, final_qty: int) -> list[ReorderEvent]:
    events: list[ReorderEvent] = []
    seen: set[tuple[str, int, dt.date | None]] = set()
    for col_idx in range(1, 50):
        value = ws.cell(24, col_idx).value
        if not isinstance(value, str) or "리오더" not in value:
            continue
        for segment in re.split(r"(?=[12]차\s*리오더)", value):
            order_match = re.search(r"([12])차\s*리오더", segment)
            qty_match = re.search(r"([\d,]+)\s*(?:PCS|CS)", segment)
            dates = re.findall(r"(?<!\d)(\d{1,2}/\d{1,2})(?!\d)", segment)
            if not order_match or not qty_match:
                continue
            event = ReorderEvent(
                f"{order_match.group(1)}차 리오더",
                int(qty_match.group(1).replace(",", "")),
                as_date(dates[-1] if dates else None),
            )
            key = (event.label, event.qty, event.date)
            if key not in seen:
                seen.add(key)
                events.append(event)
    events = sorted(events, key=lambda event: (event.label, event.date or dt.date.max, event.qty))
    reorder_stage = str(ws["J13"].value or "1차 리오더")
    if "2차" in reorder_stage:
        first = next((event for event in events if event.label.startswith("1차")), None)
        second = next((event for event in events if event.label.startswith("2차")), None)
        result: list[ReorderEvent] = []
        if first:
            result.append(first)
        result.append(ReorderEvent("2차 리오더", final_qty, second.date if second else None))
        return result
    first_date = next((event.date for event in events if event.label.startswith("1차") and event.date), None)
    return [ReorderEvent("1차 리오더", final_qty, first_date)] if final_qty else []


def load_style(ws) -> StyleData:
    inbound = as_int(ws["I3"].value)
    sales = as_int(ws["I4"].value)
    reorder_request = as_int(ws["I13"].value)
    weeks: list[dt.date] = []
    labels: list[str] = []
    weekly_sales: list[float] = []
    cum_sales: list[float] = []
    rates_pct: list[float] = []

    for idx in range(13, 50):  # M:AW
        col = col_name(idx)
        week = as_date(ws[f"{col}26"].value)
        if not week:
            continue
        cumulative = as_float(ws[f"{col}29"].value)
        weeks.append(week)
        labels.append(f"{week.month}/{week.day}")
        weekly_sales.append(as_float(ws[f"{col}27"].value))
        cum_sales.append(cumulative)
        rates_pct.append((cumulative / inbound * 100) if inbound else 0)

    review_date = as_date(ws["B16"].value) or dt.date(2026, 6, 4)
    return StyleData(
        sheet=ws.title,
        sku=str(ws["A5"].value or f"WA2602{ws.title.replace(' ', '')}"),
        review_date=review_date,
        inbound=inbound,
        sales=sales,
        inventory=inbound - sales,
        reorder_request=reorder_request,
        reorders=parse_reorders(ws, reorder_request),
        weeks=weeks,
        labels=labels,
        weekly_sales=weekly_sales,
        cum_sales=cum_sales,
        rates_pct=rates_pct,
    )


def stockout_interval(data: StyleData) -> tuple[str, int | None]:
    prev_label = "초기"
    for idx, rate in enumerate(data.rates_pct):
        if rate >= 100:
            return f"{prev_label}~{data.labels[idx]}", idx
        prev_label = data.labels[idx]
    return "표시 구간 내 초도 재고 완전 소진 예상 없음", None


def stockout_status(data: StyleData, stockout_idx: int | None) -> str:
    if stockout_idx is None:
        return "소진 예상"
    return "소진 완료" if data.weeks[stockout_idx] < data.review_date else "소진 예상"


def supply_series(data: StyleData) -> list[int]:
    supply: list[int] = []
    running = data.inbound
    applied: set[int] = set()
    for week in data.weeks:
        for idx, event in enumerate(data.reorders):
            if idx not in applied and event.date and event.date <= week:
                running += event.qty
                applied.add(idx)
        supply.append(running)
    return supply


def card(fig, x: float, y: float, w: float, h: float, title: str, value: str, note: str) -> None:
    fig.patches.append(
        plt.Rectangle((x, y), w, h, transform=fig.transFigure, facecolor="#fffdf8", edgecolor="#d8c9ad", lw=1)
    )
    fig.text(x + 0.02, y + h - 0.027, title, fontsize=10.5, color="#655640", weight="bold")
    fig.text(x + 0.02, y + 0.027, value, fontsize=16, color="#111111", weight="bold")
    if note:
        fig.text(x + 0.02, y + 0.008, note, fontsize=8.5, color="#8a7a61")


def legend(fig) -> None:
    x0, y0, w, h = 0.08, 0.07, 0.84, 0.09
    fig.patches.append(
        plt.Rectangle((x0, y0), w, h, transform=fig.transFigure, facecolor="#fffdf8", edgecolor="#d8c9ad", lw=1)
    )
    items = [
        ("#d84b3f", "-", "o", "초도 잔여 재고"),
        ("#202020", "-", None, "누적 판매량"),
        ("#1f789b", (0, (5, 4)), None, "누적 공급량"),
        ("#9bd5ce", "-", None, "주차별 판매"),
        ("#8f52ff", "-.", None, "누적 소진율"),
    ]
    positions = [(0, 0), (1, 0), (2, 0), (0, 1), (1, 1)]
    for (gx, gy), (color, style, marker, label) in zip(positions, items):
        x = x0 + 0.04 + gx * 0.27
        y = y0 + 0.058 - gy * 0.035
        fig.lines.append(
            plt.Line2D(
                [x, x + 0.04],
                [y, y],
                transform=fig.transFigure,
                color=color,
                linestyle=style,
                marker=marker,
                markersize=4,
                lw=2,
            )
        )
        fig.text(x + 0.055, y - 0.008, label, fontsize=9, color="#4e4537")


def draw_page(data: StyleData, output_png: Path | None = None):
    fig = plt.figure(figsize=(16, 9), facecolor="#f4efe4")
    fig.text(0.05, 0.92, f"{data.sku} 리오더 진행 근거", fontsize=25, weight="bold", color="#111111")
    fig.text(0.05, 0.875, "L25:AW29 예상 판매 흐름 기준", fontsize=13, color="#6d665c")

    badge = "2차 리오더" if any(event.label.startswith("2차") for event in data.reorders) else "1차 리오더"
    fig.patches.append(
        plt.Rectangle((0.765, 0.895), 0.075, 0.037, transform=fig.transFigure, facecolor="#0f6374", edgecolor="#0f6374")
    )
    fig.text(0.8025, 0.908, badge, fontsize=12, color="white", weight="bold", ha="center")

    final_label = data.labels[-1] if data.labels else "-"
    final_sales = data.cum_sales[-1] if data.cum_sales else 0
    final_rate = data.rates_pct[-1] if data.rates_pct else 0
    card_count = 3 + len(data.reorders)
    gap = 0.016
    start_x = 0.065
    card_w = (0.87 - gap * (card_count - 1)) / card_count
    cards = [
        ("1차 입고량", f"{data.inbound:,}장", "I3 기준"),
        ("현재 누적 판매", f"{data.sales:,}장", f"검토일 {data.review_date:%Y-%m-%d}"),
        ("현재 재고", f"{data.inventory:,}장", "I3 - I4"),
    ]
    for event in data.reorders:
        cards.append((event.label, f"{event.qty:,}장", f"{event.date.month}/{event.date.day} 입고" if event.date else ""))
    for idx, item in enumerate(cards):
        card(fig, start_x + idx * (card_w + gap), 0.72, card_w, 0.078, *item)

    interval, stockout_idx = stockout_interval(data)
    depletion_status = stockout_status(data, stockout_idx)
    fig.text(0.08, 0.685, "초도 재고 소진 및 리오더 공급 흐름", fontsize=17, weight="bold", color="#111111")
    if stockout_idx is None:
        fig.text(0.08, 0.665, interval, fontsize=11, color="#5d5144")
    else:
        fig.text(0.08, 0.665, f"{interval} 사이 초도 재고 {depletion_status}", fontsize=11, color="#5d5144")

    ax = fig.add_axes([0.08, 0.29, 0.82, 0.345], facecolor="#fffdf8")
    x = list(range(len(data.weeks)))
    remaining = [max(data.inbound - value, 0) for value in data.cum_sales]
    supply = supply_series(data)

    ax.plot(x, remaining, color="#d84b3f", marker="o", markersize=3.4, lw=2)
    ax.plot(x, data.cum_sales, color="#202020", lw=2.2)
    ax.plot(x, supply, color="#1f789b", lw=2, linestyle=(0, (5, 4)))
    ax.plot(x, data.weekly_sales, color="#9bd5ce", lw=2)
    y_top = ax.get_ylim()[1]

    ax2 = ax.twinx()
    ax2.plot(x, data.rates_pct, color="#8f52ff", lw=2.2, linestyle="-.")
    review_rate = data.sales / data.inbound * 100 if data.inbound else 0

    review_idx_float: float | None = None
    if data.weeks:
        if data.review_date <= data.weeks[0]:
            review_idx_float = 0.0
        elif data.review_date >= data.weeks[-1]:
            review_idx_float = float(len(data.weeks) - 1)
        else:
            for idx in range(1, len(data.weeks)):
                prev_week = data.weeks[idx - 1]
                week = data.weeks[idx]
                if prev_week <= data.review_date <= week:
                    span = (week - prev_week).days or 1
                    review_idx_float = (idx - 1) + (data.review_date - prev_week).days / span
                    break
    if review_idx_float is not None:
        ax.axvline(review_idx_float, color="#8a8378", lw=1.2, linestyle=(0, (2, 3)))
        ax.annotate(
            f"검토일 {data.review_date.month}/{data.review_date.day}\n판매율 {review_rate:.1f}%",
            xy=(review_idx_float, data.sales),
            xytext=(review_idx_float + 0.45, y_top * 0.18),
            fontsize=8.5,
            color="#6c4cc2",
            bbox=dict(facecolor="#fff7fb", edgecolor="#c7a7f2", lw=0.8, pad=3),
        )

    for reorder_idx, event in enumerate(data.reorders):
        if not event.date:
            continue
        event_idx = next((idx for idx, week in enumerate(data.weeks) if week >= event.date), None)
        if event_idx is None:
            continue
        ax.axvline(event_idx, color="#e88434", lw=1.2, linestyle=":")
        label_y = y_top * (0.82 - reorder_idx * 0.13)
        ax.annotate(
            f"{event.date.month}/{event.date.day} {event.label}\n+{event.qty:,}장",
            xy=(event_idx, supply[event_idx]),
            xytext=(event_idx + 0.2, label_y),
            fontsize=8,
            color="#a65f15",
            arrowprops=dict(arrowstyle="-", color="#a65f15", lw=0.7),
            bbox=dict(facecolor="#fff8eb", edgecolor="#e3b46f", lw=0.8, pad=3),
        )

    if stockout_idx is not None:
        left = max(stockout_idx - 1, 0)
        ax.axvspan(left, stockout_idx, color="#e9b3a3", alpha=0.25)
        ax.text(
            max(stockout_idx - 3.4, 0),
            y_top * 0.72,
            f"{depletion_status}\n{interval}",
            fontsize=8.5,
            color="#a14b42",
            bbox=dict(facecolor="#fff4ee", edgecolor="#d9a092", lw=0.8, pad=4),
        )

    ax.annotate(
        f"{final_label} 예상\n판매 {final_sales:,.0f}장\n소진율 {final_rate:.1f}%",
        xy=(len(data.weeks) - 1, data.cum_sales[-1] if data.cum_sales else 0),
        xytext=(0.78, 0.70),
        xycoords="data",
        textcoords=ax.transAxes,
        fontsize=8.5,
        color="#5f503d",
        ha="left",
        va="top",
        arrowprops=dict(arrowstyle="->", color="#8a7a61", lw=0.8),
        bbox=dict(facecolor="#fff8eb", edgecolor="#d8c9ad", lw=0.8, pad=4),
    )

    ax.set_xticks(x[::2])
    ax.set_xticklabels([data.labels[idx] for idx in x[::2]], fontsize=8)
    ax.set_ylabel("수량", fontsize=10)
    ax2.set_ylabel("누적 소진율", fontsize=10)
    ax.grid(True, axis="y", color="#e4d7c6", lw=0.7)
    ax.spines[["top", "right"]].set_visible(False)
    ax2.spines[["top", "left"]].set_visible(False)
    ax2.set_ylim(bottom=0)

    legend(fig)
    fig.text(
        0.05,
        0.028,
        "데이터 기준: 260604 Workbook. 누적 소진율은 I3(1차 입고량)만 100% 기준으로 계산하며, 소진 시점은 L25:AW29 판매 흐름 기준.",
        fontsize=8.5,
        color="#9a9185",
    )
    if output_png:
        fig.savefig(output_png, dpi=160, facecolor=fig.get_facecolor())
    return fig


def find_workbook() -> Path:
    matches = [p for d in Path(".").iterdir() if d.is_dir() for p in d.glob("*.xlsx") if "260604" in p.name]
    if not matches:
        raise SystemExit("260604 workbook not found")
    return max(matches, key=lambda path: path.stat().st_mtime)


def workbook_copy(path: Path) -> Path:
    temp_dir = Path(tempfile.mkdtemp(prefix="reorder_260604_"))
    copy_path = temp_dir / path.name
    try:
        import win32com.client  # type: ignore

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(path.resolve()), ReadOnly=True)
        workbook.SaveCopyAs(str(copy_path.resolve()))
        workbook.Close(False)
        excel.Quit()
    except Exception:
        shutil.copy2(path, copy_path)
    return copy_path


def main() -> None:
    set_font()
    workbook = workbook_copy(find_workbook())
    out_dir = Path("output/reorder_analysis_basis_pages_260604")
    out_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(workbook, data_only=True, read_only=True)
    data = [load_style(wb[target]) for target in TARGETS]

    pdf_path = out_dir / "reorder_analysis_ST32_ST27_ST16_ST34_ST31_ST24_260604.pdf"
    with PdfPages(pdf_path) as pdf:
        for index, item in enumerate(data, 1):
            png_path = out_dir / f"_preview_260604_page{index:02d}_{item.sheet.replace(' ', '_')}.png"
            fig = draw_page(item, png_path)
            pdf.savefig(fig, facecolor=fig.get_facecolor())
            plt.close(fig)

    summary = out_dir / "summary.tsv"
    summary.write_text(
        "sheet\tsku\treview_date\tinbound\tsales\tinventory\treorders\n"
        + "\n".join(
            f"{item.sheet}\t{item.sku}\t{item.review_date:%Y-%m-%d}\t{item.inbound}\t{item.sales}\t{item.inventory}\t"
            + "; ".join(f"{event.label} {event.qty} {event.date or ''}" for event in item.reorders)
            for item in data
        ),
        encoding="utf-8",
    )
    print(pdf_path)
    print(summary)


if __name__ == "__main__":
    main()
