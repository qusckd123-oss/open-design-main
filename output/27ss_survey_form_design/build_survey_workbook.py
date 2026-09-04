import json
import os
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


def add_table(ws):
    if ws.max_row <= 1 or ws.max_column <= 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    name = re.sub("[^A-Za-z0-9_]", "_", ws.title)[:25] + "_tbl"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


src = os.environ["XLSX_PATH"]
out_dir = Path("output/27ss_survey_form_design")
out_dir.mkdir(parents=True, exist_ok=True)
out = out_dir / "WACKYWILLY_27SS_품평_설문폼_설계.xlsx"

wb_src = load_workbook(src, data_only=True, read_only=False)
styles = []
for sheet, line in [("27SS UNI", "UNISEX"), ("27SS WOMAN", "WOMEN")]:
    ws = wb_src[sheet]
    for row in range(5, ws.max_row + 1):
        style_no = ws.cell(row, 6).value
        item_name = ws.cell(row, 8).value
        if not style_no or not item_name or str(style_no).startswith("#"):
            continue
        style_no = str(style_no).strip()
        item_name = str(item_name).strip()
        image_val = ws.cell(row, 7).value
        match = re.search(r"\d{4}([A-Z]+)\d+", style_no)
        category = match.group(1) if match else ""
        styles.append(
            OrderedDict(
                [
                    ("라인", line),
                    ("카테고리", category),
                    ("STYLE NO.", style_no),
                    ("MAIN NUMBER", ws.cell(row, 3).value or ""),
                    ("ITEM NAME", item_name),
                    ("IMAGE", "" if image_val in (None, "#VALUE!") else image_val),
                    ("이미지_URL/파일경로_입력", ""),
                    ("RRP(KRW)", ws.cell(row, 19).value or ""),
                    ("FIT", "" if ws.cell(row, 16).value in (None, 0) else ws.cell(row, 16).value),
                    ("디자인 설명", ws.cell(row, 17).value or ""),
                    ("폼 섹션명", f"{line} / {category}"),
                    ("폼 표시 제목", f"[{line}/{category}] {style_no} | {item_name}"),
                ]
            )
        )

seen = set()
unique = []
for style in styles:
    key = (style["라인"], style["카테고리"], style["STYLE NO."])
    if key not in seen:
        seen.add(key)
        unique.append(style)
styles = unique

categories = []
seen = set()
for style in styles:
    key = (style["라인"], style["카테고리"])
    if key not in seen:
        seen.add(key)
        categories.append(key)

wb = Workbook()
brand_fill = PatternFill("solid", fgColor="111827")
head_fill = PatternFill("solid", fgColor="D9EAF7")
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws = wb.active
ws.title = "00_Guide"
ws["A1"] = "WACKYWILLY 27SS 품평 설문폼 설계"
ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
ws["A1"].fill = brand_fill
ws.merge_cells("A1:H1")
for row in [
    ["목적", "매장 매니저 및 사업부 임직원 의견을 품번별/카테고리별로 동일 기준 수집"],
    ["추천 운영", "QR 1개로 설문 링크 진입 → 응답자 정보 입력 → 카테고리별 섹션 이동 → 품번별 객관식 평가 + 선택 주관 의견"],
    ["추천 폼 도구", "Microsoft Forms 또는 Google Forms. 이미지 표시가 중요하면 Google Forms가 품번별 이미지+객관식 그리드 구성에 유리함."],
    ["이미지 처리", "원본 라인시트에는 실제 삽입 이미지가 없고 IMAGE 값이 #VALUE!로 표시됨. 이미지 URL 또는 파일경로를 01_Style_Master의 이미지_URL/파일경로_입력에 추가 후 폼에 삽입 필요."],
    ["기본 평가 범위", "5점 척도: 1 매우 낮음 / 2 낮음 / 3 보통 / 4 높음 / 5 매우 높음"],
    ["전개 추천 범위", "제외 / 소량 테스트 / 기본 전개 / 확대 전개 / 주력 전개"],
    ["주관식 운영", "품번별 주관 의견은 선택 입력, 카테고리 종합 평가는 필수 입력 권장"],
]:
    ws.append(row)

ws2 = wb.create_sheet("00_Score_Guide")
for row in [
    ["평가항목", "문항", "선택지/범위", "점수 해석", "필수여부"],
    ["진행 적합도", "이 품번을 27SS에 전개할 가치가 있습니까?", "1 매우 낮음, 2 낮음, 3 보통, 4 높음, 5 매우 높음", "4~5점은 진행 추천, 1~2점은 드랍/재검토 후보", "필수"],
    ["예상 판매성", "매장/채널에서 판매 전환 가능성이 높습니까?", "1~5점", "매장 매니저 의견 핵심 지표", "필수"],
    ["디자인/컬러 매력도", "디자인과 컬러가 고객에게 매력적으로 보입니까?", "1~5점", "상품 외관/컬러 반응 지표", "필수"],
    ["가격 적정성", "예상 판매가 기준 가격 저항이 낮습니까?", "1 매우 비쌈~5 매우 적정", "낮을수록 가격 저항 높음", "필수"],
    ["브랜드/고객 적합도", "와키윌리 고객과 매장 무드에 잘 맞습니까?", "1~5점", "브랜드 핏 판단", "필수"],
    ["차별성/신선도", "기존 상품 대비 새로움이나 차별성이 있습니까?", "1~5점", "중복/식상함 판단", "필수"],
    ["전개 추천", "전개 수량/범위를 어떻게 추천합니까?", "제외, 소량 테스트, 기본 전개, 확대 전개, 주력 전개", "수량 배분 의사결정용", "필수"],
    ["품번별 주관 의견", "좋은 점/우려점/수정 의견이 있으면 작성해 주세요.", "자유 입력", "선택 입력", "선택"],
]:
    ws2.append(row)

ws3 = wb.create_sheet("01_Style_Master")
ws3.append(list(styles[0].keys()))
for style in styles:
    ws3.append(list(style.values()))

questions = [
    ("Q1_진행적합도", "이 품번을 27SS에 전개할 가치가 있습니까?", "객관식 5점", "1 매우 낮음 / 2 낮음 / 3 보통 / 4 높음 / 5 매우 높음", "필수"),
    ("Q2_예상판매성", "매장/채널에서 판매 전환 가능성이 높습니까?", "객관식 5점", "1 매우 낮음 / 2 낮음 / 3 보통 / 4 높음 / 5 매우 높음", "필수"),
    ("Q3_디자인컬러", "디자인과 컬러가 고객에게 매력적으로 보입니까?", "객관식 5점", "1 매우 낮음 / 2 낮음 / 3 보통 / 4 높음 / 5 매우 높음", "필수"),
    ("Q4_가격적정성", "예상 판매가 기준 가격 저항이 낮습니까?", "객관식 5점", "1 매우 비쌈 / 2 다소 비쌈 / 3 보통 / 4 적정 / 5 매우 적정", "필수"),
    ("Q5_브랜드고객핏", "와키윌리 고객과 매장 무드에 잘 맞습니까?", "객관식 5점", "1 매우 낮음 / 2 낮음 / 3 보통 / 4 높음 / 5 매우 높음", "필수"),
    ("Q6_차별성신선도", "기존 상품 대비 새로움이나 차별성이 있습니까?", "객관식 5점", "1 매우 낮음 / 2 낮음 / 3 보통 / 4 높음 / 5 매우 높음", "필수"),
    ("Q7_전개추천", "전개 수량/범위를 어떻게 추천합니까?", "객관식", "제외 / 소량 테스트 / 기본 전개 / 확대 전개 / 주력 전개", "필수"),
    ("Q8_품번의견", "좋은 점/우려점/수정 의견이 있으면 작성해 주세요.", "주관식", "자유 입력", "선택"),
]

ws4 = wb.create_sheet("02_Style_Questions")
ws4.append(["라인", "카테고리", "STYLE NO.", "ITEM NAME", "이미지_URL/파일경로", "폼 표시 제목", "문항코드", "문항", "문항유형", "선택지", "필수여부"])
for style in styles:
    for q in questions:
        ws4.append([style["라인"], style["카테고리"], style["STYLE NO."], style["ITEM NAME"], style["이미지_URL/파일경로_입력"], style["폼 표시 제목"], *q])

ws5 = wb.create_sheet("03_Category_Questions")
ws5.append(["라인", "카테고리", "문항코드", "문항", "문항유형", "필수여부", "운영의도"])
category_questions = [
    ("C1_카테고리강점", "이 카테고리에서 가장 강하다고 느껴지는 상품/디테일은 무엇입니까?", "주관식", "필수", "강점 상품/요소 도출"),
    ("C2_카테고리우려", "판매가 우려되는 상품군, 가격대, 디자인 요소는 무엇입니까?", "주관식", "필수", "리스크 도출"),
    ("C3_누락상품", "현재 구성에서 빠졌다고 느껴지는 아이템/컬러/핏이 있습니까?", "주관식", "필수", "구성 보완"),
    ("C4_가격반응", "카테고리 전체 가격대에 대한 고객 반응을 어떻게 예상합니까?", "주관식", "필수", "가격 전략"),
    ("C5_매장전개", "매장 전개 시 우선 보여줘야 할 상품과 연출 방향은 무엇입니까?", "주관식", "필수", "VM/매장 운영"),
    ("C6_최종의견", "기타 자유 의견을 작성해 주세요.", "주관식", "선택", "추가 의견"),
]
for line, category in categories:
    for q in category_questions:
        ws5.append([line, category, *q])

ws6 = wb.create_sheet("04_Response_Longform")
ws6.append(["응답ID", "응답일시", "응답자명", "소속/매장", "응답자구분", "라인", "카테고리", "STYLE NO.", "ITEM NAME", "진행적합도", "예상판매성", "디자인컬러", "가격적정성", "브랜드고객핏", "차별성신선도", "전개추천", "품번별주관의견"])
for style in styles:
    ws6.append(["", "", "", "", "", style["라인"], style["카테고리"], style["STYLE NO."], style["ITEM NAME"], "", "", "", "", "", "", "", ""])

ws7 = wb.create_sheet("05_QR_Guide")
for row in [
    ["항목", "내용"],
    ["QR 링크", "폼 생성 후 실제 URL을 여기에 붙여넣기"],
    ["게시 위치", "품평장 입구, 카테고리별 존, 샘플랙 옆"],
    ["권장 안내 문구", "27SS 품평 설문: QR 접속 후 소속/성명 입력, 카테고리별 품번 평가와 종합 의견을 남겨주세요."],
    ["응답 소요", "품번 수가 많으므로 카테고리별 응답 시간을 분리 운영 권장"],
    ["이미지 운영", "폼에는 STYLE NO. / ITEM NAME과 함께 이미지 업로드 또는 이미지 URL 삽입 필요"],
    ["응답 마감", "품평 당일 종료 전 1차 마감, 필요 시 익일 오전까지 추가 응답"],
]:
    ws7.append(row)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = brand_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.row != 1:
                cell.font = Font(name="Arial", size=10)
    for col in range(1, min(ws.max_column, 20) + 1):
        max_len = 10
        for cell in ws[get_column_letter(col)][:100]:
            if cell.value is not None:
                max_len = max(max_len, min(60, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col)].width = max_len
    if ws.title != "00_Guide":
        add_table(ws)

score_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
rec_dv = DataValidation(type="list", formula1='"제외,소량 테스트,기본 전개,확대 전개,주력 전개"', allow_blank=True)
ws6.add_data_validation(score_dv)
ws6.add_data_validation(rec_dv)
for col in ["J", "K", "L", "M", "N", "O"]:
    score_dv.add(f"{col}2:{col}{max(2, ws6.max_row)}")
rec_dv.add(f"P2:P{max(2, ws6.max_row)}")

wb.save(out)
print(out)
print(json.dumps({"styles": len(styles), "categories": len(categories)}, ensure_ascii=False))
