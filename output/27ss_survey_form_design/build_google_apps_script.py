import json
from pathlib import Path

from openpyxl import load_workbook


base = Path("output/27ss_survey_form_design")
xlsx = base / "WACKYWILLY_27SS_품평_설문폼_설계.xlsx"
out = base / "create_27ss_wackywilly_review_form.gs"

wb = load_workbook(xlsx, data_only=True)
ws = wb["01_Style_Master"]
headers = [c.value for c in ws[1]]
styles = []
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if d.get("STYLE NO.") and d.get("ITEM NAME"):
        styles.append(
            {
                "line": d.get("라인") or "",
                "category": d.get("카테고리") or "",
                "styleNo": d.get("STYLE NO.") or "",
                "itemName": d.get("ITEM NAME") or "",
                "imageUrl": d.get("이미지_URL/파일경로_입력") or "",
                "rrp": d.get("RRP(KRW)") or "",
                "fit": d.get("FIT") or "",
                "description": d.get("디자인 설명") or "",
            }
        )

script = f"""/**
 * WACKYWILLY 27SS 품평 Google Form 자동 생성 스크립트
 *
 * 사용 전 확인:
 * 1. Google Drive에서 새 Google Apps Script 프로젝트 생성
 * 2. 이 파일 전체를 Code.gs에 붙여넣기
 * 3. 상단 실행 함수에서 createWackyWilly27SSReviewForm 선택
 * 4. Run 실행 후 권한 승인
 * 5. 실행 로그에서 EDIT_URL / RESPONDENT_URL 확인
 *
 * 이미지:
 * - 현재 라인시트 IMAGE 컬럼은 #VALUE! 상태라 imageUrl이 비어 있습니다.
 * - 품번별 이미지를 넣으려면 STYLE_DATA의 imageUrl에 공개 접근 가능한 이미지 URL을 입력하세요.
 */

const FORM_TITLE = 'WACKYWILLY 27SS 품평 설문';
const FORM_DESCRIPTION = [
  '매장 매니저 및 사업부 임직원 대상 27SS 품평 설문입니다.',
  '각 품번별 객관식 평가는 1~5점 기준입니다.',
  '1 매우 낮음 / 2 낮음 / 3 보통 / 4 높음 / 5 매우 높음',
  '전개 추천은 제외 / 소량 테스트 / 기본 전개 / 확대 전개 / 주력 전개 중 선택해 주세요.',
  '품번별 주관 의견은 선택 입력, 카테고리 종합 의견은 필수 입력입니다.'
].join('\\n');

const STYLE_DATA = {json.dumps(styles, ensure_ascii=False, indent=2)};

function createWackyWilly27SSReviewForm() {{
  const form = FormApp.create(FORM_TITLE);
  form.setDescription(FORM_DESCRIPTION);
  form.setCollectEmail(false);
  form.setAllowResponseEdits(false);
  form.setShowLinkToRespondAgain(false);
  form.setConfirmationMessage('응답이 저장되었습니다. 감사합니다.');

  addRespondentQuestions_(form);

  const grouped = groupByLineCategory_(STYLE_DATA);
  Object.keys(grouped).forEach(function(sectionKey) {{
    const sectionStyles = grouped[sectionKey];
    const first = sectionStyles[0];
    form.addPageBreakItem()
      .setTitle(first.line + ' / ' + first.category)
      .setHelpText('해당 카테고리 품번별 평가 후 마지막에 카테고리 종합 의견을 작성해 주세요.');

    sectionStyles.forEach(function(style) {{
      addStyleQuestionSet_(form, style);
    }});

    addCategoryQuestions_(form, first.line, first.category);
  }});

  Logger.log('EDIT_URL: ' + form.getEditUrl());
  Logger.log('RESPONDENT_URL: ' + form.getPublishedUrl());
  Logger.log('TOTAL_STYLES: ' + STYLE_DATA.length);
  Logger.log('TOTAL_SECTIONS: ' + Object.keys(grouped).length);
}}

function addRespondentQuestions_(form) {{
  form.addSectionHeaderItem()
    .setTitle('응답자 정보')
    .setHelpText('품평 응답 취합을 위한 기본 정보입니다.');

  form.addTextItem()
    .setTitle('성명')
    .setRequired(true);

  form.addTextItem()
    .setTitle('소속 / 매장명')
    .setRequired(true);

  form.addListItem()
    .setTitle('응답자 구분')
    .setChoiceValues(['매장 매니저', '영업/영업기획', '상품기획', '디자인', '소싱/소재', '마케팅/온라인/VM', '사업부 임직원', '기타'])
    .setRequired(true);
}}

function addStyleQuestionSet_(form, style) {{
  const title = '[' + style.line + '/' + style.category + '] ' + style.styleNo + ' | ' + style.itemName;
  const help = [
    style.rrp ? 'RRP: ' + style.rrp : '',
    style.fit ? 'FIT: ' + style.fit : '',
    style.description ? '설명: ' + style.description : ''
  ].filter(Boolean).join('\\n');

  form.addSectionHeaderItem()
    .setTitle(title)
    .setHelpText(help);

  if (style.imageUrl) {{
    try {{
      const blob = UrlFetchApp.fetch(style.imageUrl).getBlob();
      form.addImageItem()
        .setTitle(style.styleNo + ' image')
        .setImage(blob);
    }} catch (err) {{
      form.addSectionHeaderItem()
        .setTitle('이미지 로드 실패: ' + style.styleNo)
        .setHelpText(String(err));
    }}
  }}

  addScale_(form, style.styleNo + ' / 진행 적합도', '이 품번을 27SS에 전개할 가치가 있습니까?');
  addScale_(form, style.styleNo + ' / 예상 판매성', '매장/채널에서 판매 전환 가능성이 높습니까?');
  addScale_(form, style.styleNo + ' / 디자인·컬러 매력도', '디자인과 컬러가 고객에게 매력적으로 보입니까?');
  addScale_(form, style.styleNo + ' / 가격 적정성', '예상 판매가 기준 가격 저항이 낮습니까?');
  addScale_(form, style.styleNo + ' / 브랜드·고객 적합도', '와키윌리 고객과 매장 무드에 잘 맞습니까?');
  addScale_(form, style.styleNo + ' / 차별성·신선도', '기존 상품 대비 새로움이나 차별성이 있습니까?');

  form.addListItem()
    .setTitle(style.styleNo + ' / 전개 추천')
    .setHelpText('해당 품번의 권장 전개 범위를 선택해 주세요.')
    .setChoiceValues(['제외', '소량 테스트', '기본 전개', '확대 전개', '주력 전개'])
    .setRequired(true);

  form.addParagraphTextItem()
    .setTitle(style.styleNo + ' / 품번별 주관 의견')
    .setHelpText('좋은 점, 우려점, 수정 의견이 있으면 작성해 주세요. 선택 입력입니다.')
    .setRequired(false);
}}

function addScale_(form, title, helpText) {{
  form.addScaleItem()
    .setTitle(title)
    .setHelpText(helpText)
    .setBounds(1, 5)
    .setLabels('매우 낮음', '매우 높음')
    .setRequired(true);
}}

function addCategoryQuestions_(form, line, category) {{
  const prefix = '[' + line + '/' + category + '] ';
  form.addSectionHeaderItem()
    .setTitle(prefix + '카테고리 종합 평가')
    .setHelpText('해당 카테고리 전체에 대한 종합 의견입니다.');

  form.addParagraphTextItem().setTitle(prefix + '가장 강하다고 느껴지는 상품/디테일은 무엇입니까?').setRequired(true);
  form.addParagraphTextItem().setTitle(prefix + '판매가 우려되는 상품군, 가격대, 디자인 요소는 무엇입니까?').setRequired(true);
  form.addParagraphTextItem().setTitle(prefix + '현재 구성에서 빠졌다고 느껴지는 아이템/컬러/핏이 있습니까?').setRequired(true);
  form.addParagraphTextItem().setTitle(prefix + '카테고리 전체 가격대에 대한 고객 반응을 어떻게 예상합니까?').setRequired(true);
  form.addParagraphTextItem().setTitle(prefix + '매장 전개 시 우선 보여줘야 할 상품과 연출 방향은 무엇입니까?').setRequired(true);
  form.addParagraphTextItem().setTitle(prefix + '기타 자유 의견').setRequired(false);
}}

function groupByLineCategory_(rows) {{
  const grouped = {{}};
  rows.forEach(function(row) {{
    const key = row.line + ' / ' + row.category;
    if (!grouped[key]) grouped[key] = [];
    grouped[key].push(row);
  }});
  return grouped;
}}
"""

out.write_text(script, encoding="utf-8")
print(out)
print(json.dumps({"styles": len(styles)}, ensure_ascii=False))
