import base64
import html
import re
from io import BytesIO
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageOps

OUT_DIR = Path("\ucd94\uac00 \ub370\uc774\ud130") / "\ud310\ub9e4 \ub370\uc774\ud130" / "260810~260816"
TREND = OUT_DIR / "\uc640\ud0a4\uc70c\ub9ac_26SS \uc804\uc0c1\ud488 \ud310\ub9e4\ucd94\uc774_260816.xlsx"
HTML_OUT = OUT_DIR / "weekly_review_teams_rich_copy_260810_260816_corrected.html"
MD_OUT = OUT_DIR / "weekly_review_teams_copy_260810_260816_corrected.md"


def data_uri(img, size=(420, 420)):
    img = ImageOps.exif_transpose(img).convert("RGB")
    img.thumbnail(size, Image.Resampling.LANCZOS)
    canvas = Image.new("RGB", size, "white")
    canvas.paste(img, ((size[0] - img.width) // 2, (size[1] - img.height) // 2))
    buf = BytesIO()
    canvas.save(buf, "JPEG", quality=88, optimize=True)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


def excel_img(code):
    wb = load_workbook(TREND, read_only=False, data_only=True)
    try:
        candidates = []
        for ws in wb.worksheets:
            hit_rows = []
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and code in cell.value:
                        hit_rows.append(cell.row)
            if not hit_rows:
                continue
            for img in getattr(ws, "_images", []):
                try:
                    r = img.anchor._from.row + 1
                    c = img.anchor._from.col + 1
                    dist = min(abs(r - h) for h in hit_rows)
                    if dist <= 16:
                        candidates.append((dist, abs(c - 7), img))
                except Exception:
                    pass
        if not candidates:
            return None
        candidates.sort(key=lambda x: (x[0], x[1]))
        return data_uri(Image.open(BytesIO(candidates[0][2]._data())))
    finally:
        wb.close()


def musinsa_img(product_id):
    page = urlopen(Request(f"https://www.musinsa.com/products/{product_id}", headers={"User-Agent": "Mozilla/5.0"}), timeout=20).read().decode("utf-8", "ignore")
    m = re.search(r'property=["\']og:image["\']\s+content=["\']([^"\']+)', page) or re.search(r'content=["\']([^"\']+)["\']\s+property=["\']og:image["\']', page)
    if not m:
        return None
    url = html.unescape(m.group(1))
    if url.startswith("//"):
        url = "https:" + url
    raw = urlopen(Request(url, headers={"User-Agent": "Mozilla/5.0"}), timeout=20).read()
    return data_uri(Image.open(BytesIO(raw)))


def direct_img(url):
    raw = urlopen(Request(url, headers={"User-Agent": "Mozilla/5.0"}), timeout=20).read()
    return data_uri(Image.open(BytesIO(raw)))


def placeholder(code):
    img = Image.new("RGB", (420, 260), "#f8f6f0")
    d = ImageDraw.Draw(img)
    try:
        big = ImageFont.truetype("arial.ttf", 32)
        small = ImageFont.truetype("arial.ttf", 18)
    except Exception:
        big = ImageFont.load_default()
        small = ImageFont.load_default()
    d.rectangle((0, 0, 419, 259), outline="#d8d1c5", width=2)
    d.text((24, 70), code, fill="#222", font=big)
    d.text((24, 128), "이미지 확인 필요", fill="#8f3f2f", font=small)
    return data_uri(img, (420, 260))


def img_for(code):
    ids = {"WA2602STC3": "6880006", "WA2602ST79": "6140637", "WA2603CD51": "7084458"}
    external = {"WA2603ST16": "https://img3.momoshop.com.tw/goodsimg/0015/099/645/spec/15099645_01_002_R.jpg?t=1774865661"}
    src = excel_img(code)
    if not src and code in ids:
        try:
            src = musinsa_img(ids[code])
        except Exception:
            src = None
    if not src and code in external:
        try:
            src = direct_img(external[code])
        except Exception:
            src = None
    return src or placeholder(code)


images = {code: img_for(code) for code in ["WA2602STC3", "WA2602ST79", "WA2603ST16", "WA2603CD51"]}

lines = [
    ("title", "■ 8월2주차 어패럴 주간 판매 요약 (26SS + 26FW)"),
    ("p", "기간: 2025.08.18 VS 2026.08.17 (ERP 기준일 2026-08-17)"),
    ("p", "전체 주간판매 6.29억 (ACC 포함)"),
    ("p", "└ 26SS 5.80억 = APP 4.97억 + ACC 0.83억"),
    ("p", "└ 26FW 0.48억(4,845만원) = APP 3,815만원 + ACC 1,030만원 — 시즌 극초반이라 절대금액이 아직 작음"),
    ("section", "1. 유니섹스/우먼스 전체 전주 판매현황"),
    ("p", "└ 26SS APP 4.97억 (전년비 -3.3% / 전주대비 -25.2%)"),
    ("p", "   유니 3.04억(전년비 -18.0%/전주대비 -23.9%) : 우먼 1.87억(전년비 +40.7%/전주대비 -27.3%) = 61.9 : 38.1"),
    ("p", "└ 26FW APP 3,815만원(0.38억) (전년비 -67.2% / 전주대비 -5.2%)"),
    ("p", "   유니 1,583만원(전년비 -77.8%/전주대비 -47.2%) : 우먼 2,232만원(전년비 -50.4%/전주대비 +129.9%) = 41.5 : 58.5"),
    ("p", "└ 기준 정리: APP 산정 시 스커트는 SR, SK는 삭스/양말류 ACC로 분류. 따라서 26FW APP에는 WA2603SKxx 양말 3PACK류를 포함하지 않음."),
    ("section", "2. 복종별 판매 현황"),
    ("p", "└ 26SS 유니 ST(유니 매출의 59%) WoW -29.9% — 먼작귀 콜라보 라인(STC1~3)은 오히려 증가세, 3PACK/베이직 라인이 크게 빠지며 전체 순감."),
    ("p", "└ 26SS 우먼 ST(우먼 매출의 45%) WoW -29.5% — WA2602ST79가 88장→9장(-90%)으로 급락, 상위권 다수 동반 하락."),
    ("p", "└ 26SS 우먼 CD 전년비 +265.7%로 우먼 최고 호조 아이템. 누계 판매율 67.4%로 우먼 평균 62.7% 상회."),
    ("p", "└ 26FW 유니는 사실상 ST 단일 아이템 중심으로 판매 발생. 신규 입고분 ST16/ST15가 증가를 견인."),
    ("p", "└ 26FW 우먼 CD 신규 런칭 96장이 우먼 WoW +129.9% 견인. 반면 우먼 ST 4개 품번은 전량 전주대비 하락."),
    ("section", "3. 복종 내 TOP 아이템/판매 채널"),
    ("issue", "※ 유니 ST — 먼작귀 콜라보 라인 반응 (전주대비 증가)"),
    ("code", "WA2602STC3 105장→209장, +551만원(+93.0%)"),
    ("p", "└ 콜라보 STC1~3 중 가장 큰 볼륨. 온라인 단독 워밍업 이후 오프라인 확장 반응 추가 확인 필요."),
    ("issue", "※ 우먼 ST — 전주대비 판매 급락 1위"),
    ("code", "WA2602ST79 88장→9장, -346만원(-90.1%)"),
    ("p", "└ 상위권 우먼 반팔 둔화의 대표 품번. 직영/백화점/면세 판매가 모두 낮아져 노출·재고·가격 조건 확인 필요."),
    ("issue", "※ 26FW 유니 ST — 신규 입고 반응"),
    ("code", "WA2603ST16 신규 입고, 이번주 58장/261만원"),
    ("p", "└ 신규 입고 직후 유니 FW ST 반응을 만든 핵심 품번. 직영점·백화점 중심 초반 판매 확인."),
    ("issue", "※ 26FW 우먼 CD — 신규 런칭 최다 판매"),
    ("code", "WA2603CD51 신규 입고, 이번주 45장/428만원"),
    ("p", "└ 우먼 CD 신규 런칭 중 최다 판매. 백화점/직영점 중심 반응이 먼저 잡힘."),
    ("section", "4. 액션/확인사항"),
    ("p", "└ 26SS ST는 먼작귀 콜라보와 3PACK/베이직 라인을 분리해서 판단 필요. 콜라보는 증가, 기존 팩/베이직은 둔화."),
    ("p", "└ 26FW는 절대금액보다 초기 반응 중심으로 볼 것. ST16, CD51은 점별 재고와 노출 유지 여부 확인."),
    ("p", "└ 다음 리뷰부터 APP/ACC 기준은 SR=스커트(APP), SK=삭스(ACC)로 고정 적용."),
    ("p", "참고 — ACC: 26SS 0.83억(전년비 -50.4%/전주대비 -17.8%) · 26FW 1,030만원(전년비 -67.8%/전주대비 -42.4%)"),
    ("p", "작성: 상품기획팀 / ERP 기준 2026-08-17 · 사진=전주대비 판매 증가 또는 이슈가 확인된 개별 품번만 표시"),
]


def esc(s):
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


html_parts, md_parts = [], []
for kind, text in lines:
    cls = {"title": "title", "section": "section", "issue": "issue", "code": "code"}.get(kind, "")
    html_parts.append(f"<p class='{cls}'>{esc(text)}</p>")
    md_parts.append(("**" + text + "**") if kind in ("title", "section", "issue") else text)
    if kind == "code":
        code = text.split()[0]
        if code in images:
            html_parts.append(f"<p><img class='product' src='{images[code]}' alt='{code}'></p>")
            md_parts.append(f"![{code}]({images[code]})")

body = "\n".join(html_parts)
doc = f"""<!doctype html>
<html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>8월2주차 어패럴 주간 판매 요약 corrected</title>
<style>
body{{margin:0;background:#eeeae3;font-family:'Malgun Gothic',Arial,sans-serif;color:#222;line-height:1.55}}
.page{{max-width:900px;margin:0 auto;padding:28px 18px 48px}}
.toolbar{{display:flex;justify-content:space-between;gap:12px;margin-bottom:12px;color:#5b554d;font-size:13px}}
button{{border:1px solid #222;background:#222;color:#fff;border-radius:6px;padding:9px 14px;font-weight:700;cursor:pointer}}
#copyArea{{background:#fff;border:1px solid #ddd6cc;padding:30px;box-shadow:0 8px 24px rgba(40,36,30,.08)}}
p{{margin:0 0 10px;white-space:pre-wrap}}
.title{{font-weight:800;font-size:22px;margin-bottom:16px}}
.section{{font-weight:800;font-size:18px;margin-top:20px;padding-top:12px;border-top:1px solid #e7e0d6}}
.issue{{font-weight:800;margin-top:16px}}
.code{{font-weight:800}}
.product{{display:block;width:210px;max-width:70%;height:auto;margin:6px 0 16px 20px;border:1px solid #ded8cf;border-radius:4px;background:#fff}}
</style></head><body><div class="page">
<div class="toolbar"><span>흰 영역을 복사하거나 버튼으로 Teams에 붙여넣기</span><button id="copyBtn">리치 복사</button></div>
<div id="copyArea" contenteditable="true">{body}</div>
</div><script>
document.getElementById('copyBtn').onclick=async()=>{{const a=document.getElementById('copyArea');try{{await navigator.clipboard.write([new ClipboardItem({{'text/html':new Blob([a.innerHTML],{{type:'text/html'}}),'text/plain':new Blob([a.innerText],{{type:'text/plain'}})}})]);}}catch(e){{const r=document.createRange();r.selectNodeContents(a);const s=window.getSelection();s.removeAllRanges();s.addRange(r);document.execCommand('copy');}}document.getElementById('copyBtn').textContent='복사 완료';}};
</script></body></html>"""

HTML_OUT.write_text(doc, encoding="utf-8")
MD_OUT.write_text("\n\n".join(md_parts), encoding="utf-8")
print(HTML_OUT)
print(MD_OUT)
print("images", len(images), "img_tags", doc.count("<img"))
