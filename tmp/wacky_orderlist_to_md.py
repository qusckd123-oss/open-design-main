from __future__ import annotations

import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


BASE = Path("와키윌리 발주리스트")
OUT = BASE / "markdown"
OUT.mkdir(parents=True, exist_ok=True)

HEADER_KEYWORDS = [
    "품번",
    "품명",
    "색상",
    "수량",
    "금액",
    "시즌",
    "택가",
    "판매가",
    "원가",
    "MARKUP",
    "진행",
    "기획",
    "생산처",
    "매입처",
    "딜리버리",
    "납기",
    "발주",
]

PREFERRED_SHEETS = [
    "상품구성안 요약",
    "상품기획안_종합",
    "선발주",
    "발주데이터",
    "ERP",
    "주간데이터",
    "발주금액 정리",
    "간단요약",
    "정리",
    "요약",
    "와키윌리 리스트",
]

SELECT_COLUMNS = [
    "시즌",
    "품번&색상",
    "품번",
    "품명",
    "색상",
    "색상명",
    "성별",
    "아이템",
    "구분",
    "기획구분",
    "진행구분",
    "진행상태",
    "최초판매가",
    "현판매가",
    "택가",
    "사전원가",
    "사전원가(V+)",
    "MARKUP",
    "배수(최)",
    "수량",
    "ERP 수량",
    "발주리스트 수량",
    "금액",
    "딜리버리일",
    "납기일자",
    "발주일자",
    "생산처명",
    "매입처명",
]


def season_from_name(name: str) -> str:
    m = re.search(r"(\d{2})(SS|FW)", name, re.I)
    return f"{m.group(1)}{m.group(2).upper()}" if m else name.replace(".xlsx", "")


def cell_text(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float):
        if math.isfinite(v) and v.is_integer():
            return str(int(v))
        return f"{v:.4f}".rstrip("0").rstrip(".")
    return str(v).replace("\n", " ").strip()


def md_escape(v) -> str:
    return cell_text(v).replace("|", "\\|")


def number(v):
    if isinstance(v, (int, float)) and math.isfinite(v):
        return float(v)
    if isinstance(v, str):
        s = v.replace(",", "").strip()
        try:
            return float(s)
        except ValueError:
            return None
    return None


def fmt_num(v) -> str:
    if v is None:
        return "-"
    if abs(v - round(v)) < 0.00001:
        return f"{int(round(v)):,}"
    return f"{v:,.1f}"


def find_header(ws):
    best = (0, None, [])
    for rix, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), 1):
        vals = [cell_text(v) for v in row]
        score = sum(1 for v in vals if any(k.lower() in v.lower() for k in HEADER_KEYWORDS))
        if score > best[0]:
            best = (score, rix, vals)
    if best[0] >= 3:
        return best[1], best[2]
    return None, []


def unique_headers(vals):
    out = []
    seen = Counter()
    for i, raw in enumerate(vals, 1):
        name = cell_text(raw) or f"COL{i}"
        seen[name] += 1
        if seen[name] > 1:
            name = f"{name}_{seen[name]}"
        out.append(name)
    return out


def chosen_indices(headers):
    picks = []
    for i, h in enumerate(headers):
        if h in SELECT_COLUMNS or any(key == h for key in SELECT_COLUMNS):
            picks.append(i)
    if not picks:
        for i, h in enumerate(headers):
            if any(key in h for key in ["품번", "품명", "색상", "시즌", "수량", "금액", "판매가", "택가", "원가", "진행", "딜리버리", "납기"]):
                picks.append(i)
    return picks[:14]


def category_from_row(headers, row):
    for key in ["아이템", "복종", "카테고리"]:
        for i, h in enumerate(headers):
            if key in h and i < len(row) and cell_text(row[i]):
                return cell_text(row[i])
    for i, h in enumerate(headers):
        if "품번" in h and i < len(row):
            code = cell_text(row[i])
            m = re.search(r"(?:WA|MG)\d{4}([A-Z]{2,3})", code)
            if m:
                return m.group(1)
    for key in ["구분"]:
        for i, h in enumerate(headers):
            if h == key and i < len(row) and cell_text(row[i]):
                return cell_text(row[i])
    return "미분류"


def preferred_col(headers, exacts, contains):
    for name in exacts:
        for i, h in enumerate(headers):
            if h == name:
                return i
    for needle in contains:
        for i, h in enumerate(headers):
            if needle in h:
                return i
    return None


def sheet_is_preferred(title: str) -> bool:
    return any(k in title for k in PREFERRED_SHEETS)


def analyze_sheet(ws):
    hrow, raw_headers = find_header(ws)
    preferred = sheet_is_preferred(ws.title)
    if not hrow and not preferred:
        return None
    headers = unique_headers(raw_headers[: ws.max_column]) if hrow else [f"COL{i}" for i in range(1, min(ws.max_column, 30) + 1)]
    indices = chosen_indices(headers)
    if not hrow:
        indices = list(range(min(ws.max_column, 10)))

    rows = []
    qty_totals = defaultdict(float)
    amt_totals = defaultdict(float)
    item_qty = defaultdict(float)
    item_amt = defaultdict(float)
    row_count = 0
    unique_styles = set()
    unique_sku = set()
    status_counts = Counter()
    quantity_col = preferred_col(headers, ["발주리스트 수량", "ERP 수량", "수량"], ["발주수량", "생산수량", "입고수량", "수량"])
    amount_col = preferred_col(headers, ["금액", "발주금액"], ["발주금액", "입고금액", "금액"])
    style_cols = [i for i, h in enumerate(headers) if h == "품번"]
    sku_cols = [i for i, h in enumerate(headers) if "품번&색상" in h]
    name_cols = [i for i, h in enumerate(headers) if h == "품명"]
    status_cols = [i for i, h in enumerate(headers) if "진행" in h]

    max_scan = min(ws.max_row, 20000)
    start = hrow + 1 if hrow else 1
    for row in ws.iter_rows(min_row=start, max_row=max_scan, values_only=True):
        vals = list(row[: len(headers)])
        if not any(cell_text(v) for v in vals):
            continue
        row_count += 1
        if len(rows) < 30:
            rows.append(vals)
        for i in style_cols:
            if i < len(vals) and cell_text(vals[i]):
                unique_styles.add(cell_text(vals[i]))
        for i in sku_cols:
            if i < len(vals) and cell_text(vals[i]):
                unique_sku.add(cell_text(vals[i]))
        for i in status_cols:
            if i < len(vals) and cell_text(vals[i]):
                status_counts[cell_text(vals[i])] += 1
        cat = category_from_row(headers, vals)
        qty = number(vals[quantity_col]) if quantity_col is not None and quantity_col < len(vals) else None
        amt = number(vals[amount_col]) if amount_col is not None and amount_col < len(vals) else None
        if qty:
            qty_totals[cat] += qty
        if amt:
            amt_totals[cat] += amt
        name = ""
        if name_cols:
            i = name_cols[0]
            if i < len(vals):
                name = cell_text(vals[i])
        style = ""
        if style_cols:
            i = style_cols[0]
            if i < len(vals):
                style = cell_text(vals[i])
        label = f"{style} {name}".strip() or "미분류"
        if qty:
            item_qty[label] += qty
        if amt:
            item_amt[label] += amt

    return {
        "title": ws.title,
        "rows": ws.max_row,
        "cols": ws.max_column,
        "header_row": hrow,
        "headers": headers,
        "indices": indices,
        "sample_rows": rows,
        "record_count": row_count,
        "unique_styles": len(unique_styles),
        "unique_sku": len(unique_sku),
        "qty_total": sum(qty_totals.values()) or None,
        "amt_total": sum(amt_totals.values()) or None,
        "qty_by_cat": sorted(qty_totals.items(), key=lambda x: x[1], reverse=True)[:12],
        "amt_by_cat": sorted(amt_totals.items(), key=lambda x: x[1], reverse=True)[:12],
        "top_qty": sorted(item_qty.items(), key=lambda x: x[1], reverse=True)[:10],
        "top_amt": sorted(item_amt.items(), key=lambda x: x[1], reverse=True)[:10],
        "status_counts": status_counts.most_common(8),
    }


def md_table(headers, rows, indices):
    if not rows or not indices:
        return "_표본 행 없음_"
    hs = [md_escape(headers[i]) for i in indices if i < len(headers)]
    lines = ["| " + " | ".join(hs) + " |", "| " + " | ".join(["---"] * len(hs)) + " |"]
    for row in rows:
        lines.append("| " + " | ".join(md_escape(row[i]) if i < len(row) else "" for i in indices[: len(hs)]) + " |")
    return "\n".join(lines)


def mini_table(title, pairs):
    if not pairs:
        return ""
    lines = [f"#### {title}", "| 항목 | 값 |", "| --- | ---: |"]
    for k, v in pairs:
        lines.append(f"| {md_escape(k)} | {fmt_num(v)} |")
    return "\n".join(lines)


season_summaries = []

for file in sorted(BASE.glob("*.xlsx"), key=lambda p: season_from_name(p.name)):
    season = season_from_name(file.name)
    wb = load_workbook(file, read_only=True, data_only=True)
    sheet_meta = [(ws.title, ws.max_row, ws.max_column) for ws in wb.worksheets]
    analyses = []
    for ws in wb.worksheets:
        info = analyze_sheet(ws)
        if info and (sheet_is_preferred(ws.title) or info["record_count"] >= 20):
            analyses.append(info)

    def main_score(candidate):
        title = candidate["title"]
        score = 0
        priority = [
            ("발주데이터ERP", 120),
            ("주간데이터ERP", 110),
            ("ERP", 100),
            ("완성(TTL)", 90),
            ("발주데이터", 85),
            ("선발주 리스트", 80),
            ("기획현황", 75),
            ("발주금액 정리", 70),
            ("와키윌리 리스트", 65),
            ("입고현황", 45),
            ("스타일 컬러 생산수량 기준", 55),
            ("상품기획안_종합", 40),
            ("상품구성안 요약", 30),
        ]
        for key, value in priority:
            if key in title:
                score = max(score, value)
        if title.lower().startswith("sheet"):
            score -= 100
        if candidate["qty_total"]:
            score += 10
        if candidate["unique_styles"]:
            score += 5
        return (score, candidate["record_count"])

    main = max(analyses, key=main_score) if analyses else None

    lines = [
        f"# {season} 와키윌리 발주리스트 Markdown",
        "",
        f"- 원본 파일: `{file.name}`",
        f"- 파일 크기: {file.stat().st_size:,} bytes",
        f"- 시트 수: {len(sheet_meta)}",
        "",
        "## 시트 인벤토리",
        "",
        "| 시트 | 행 | 열 |",
        "| --- | ---: | ---: |",
    ]
    for title, rows, cols in sheet_meta:
        lines.append(f"| {md_escape(title)} | {rows:,} | {cols:,} |")
    lines += ["", "## 주요 데이터 시트 분석", ""]

    for info in analyses:
        lines += [
            f"### {info['title']}",
            "",
            f"- 원본 범위: {info['rows']:,}행 x {info['cols']:,}열",
            f"- 추정 헤더 행: {info['header_row'] or '-'}",
            f"- 분석 레코드 수: {info['record_count']:,}",
            f"- 고유 품번 수: {info['unique_styles']:,}",
            f"- 고유 품번&색상 수: {info['unique_sku']:,}",
            f"- 수량 합계: {fmt_num(info['qty_total'])}",
            f"- 금액 합계: {fmt_num(info['amt_total'])}",
            "",
        ]
        if info["status_counts"]:
            lines += ["#### 진행 상태", "| 상태 | 건수 |", "| --- | ---: |"]
            for k, v in info["status_counts"]:
                lines.append(f"| {md_escape(k)} | {v:,} |")
            lines.append("")
        for block in [
            mini_table("카테고리별 수량 상위", info["qty_by_cat"]),
            mini_table("카테고리별 금액 상위", info["amt_by_cat"]),
            mini_table("상품별 수량 상위", info["top_qty"]),
            mini_table("상품별 금액 상위", info["top_amt"]),
        ]:
            if block:
                lines += [block, ""]
        lines += ["#### 대표 행", md_table(info["headers"], info["sample_rows"], info["indices"]), ""]

    if main:
        season_summaries.append(
            {
                "season": season,
                "file": file.name,
                "main_sheet": main["title"],
                "records": main["record_count"],
                "styles": main["unique_styles"],
                "sku": main["unique_sku"],
                "qty": main["qty_total"],
                "amt": main["amt_total"],
                "qty_by_cat": main["qty_by_cat"][:8],
                "top_qty": main["top_qty"][:8],
            }
        )

    (OUT / f"{season}_발주리스트_분석.md").write_text("\n".join(lines), encoding="utf-8-sig")
    wb.close()

mem = [
    "# 와키윌리 시즌별 발주리스트 메모리얼",
    "",
    "이 문서는 `와키윌리 발주리스트` 폴더의 시즌별 Excel 발주/상품기획 파일을 Markdown으로 변환하며 정리한 인덱스입니다.",
    "",
    "## 파일별 핵심 요약",
    "",
    "| 시즌 | 원본 파일 | 대표 시트 | 레코드 | 고유 품번 | 고유 품번&색상 | 수량 합계 | 금액 합계 |",
    "| --- | --- | --- | ---: | ---: | ---: | ---: | ---: |",
]
for s in season_summaries:
    mem.append(
        f"| {s['season']} | `{md_escape(s['file'])}` | {md_escape(s['main_sheet'])} | "
        f"{s['records']:,} | {s['styles']:,} | {s['sku']:,} | {fmt_num(s['qty'])} | {fmt_num(s['amt'])} |"
    )

mem += [
    "",
    "## 시즌별 카테고리 메모",
    "",
]
for s in season_summaries:
    mem += [f"### {s['season']}", "", f"- 대표 분석 시트: `{s['main_sheet']}`", "- 카테고리별 수량 상위:"]
    if s["qty_by_cat"]:
        for k, v in s["qty_by_cat"]:
            mem.append(f"  - {k}: {fmt_num(v)}")
    else:
        mem.append("  - 수량 컬럼을 안정적으로 특정하지 못함")
    mem.append("- 상품별 수량 상위:")
    if s["top_qty"]:
        for k, v in s["top_qty"]:
            mem.append(f"  - {k}: {fmt_num(v)}")
    else:
        mem.append("  - 수량 컬럼을 안정적으로 특정하지 못함")
    mem.append("")

mem += [
    "## 데이터 해석 메모",
    "",
    "- 23SS~24SS 파일은 과거 `와릿이즌/MG` 코드와 카테고리별 기획 시트가 혼재되어 있고, 일부 시트는 Excel 서식 때문에 행 수가 과대 표기됩니다.",
    "- 25SS 이후 파일은 `WA` 품번 체계, `주간데이터ERP`, `발주데이터ERP`, `상품기획안_종합`, `선발주 리스트` 계열이 중심입니다.",
    "- 26SS 파일은 `발주데이터ERP`에 ERP 수량과 발주리스트 수량 검증 컬럼이 함께 있어 실제 발주 검증용 기준 시트로 보기 좋습니다.",
    "- 26FW 파일은 `26FW 선발주 리스트`, `발주데이터`, `발주금액 정리`, `상품기획안_종합`이 함께 있어 선발주와 시즌 전체 구성 비교에 적합합니다.",
    "- 금액/수량 합계는 시트별 컬럼명이 다르고 일부 시트에 집계행이 포함될 수 있어, 세부 의사결정 전에는 해당 시즌 Markdown의 대표 행과 원본 Excel을 함께 확인해야 합니다.",
    "",
    "## 생성 파일",
    "",
]
for p in sorted(OUT.glob("*_발주리스트_분석.md")):
    mem.append(f"- `{p.name}`")

(OUT / "00_와키윌리_시즌별_발주리스트_메모리얼.md").write_text("\n".join(mem), encoding="utf-8-sig")

print(f"created {len(list(OUT.glob('*.md')))} markdown files in {OUT}")
