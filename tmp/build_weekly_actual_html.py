import base64
import html
import json
import re
from io import BytesIO
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageOps

BASE = Path("\ucd94\uac00 \ub370\uc774\ud130") / "\ud310\ub9e4 \ub370\uc774\ud130"
OUT_DIR = BASE / "260810~260816"
TREND = OUT_DIR / "\uc640\ud0a4\uc70c\ub9ac_26SS \uc804\uc0c1\ud488 \ud310\ub9e4\ucd94\uc774_260816.xlsx"
ANALYSIS = Path("tmp") / "weekly_actual_analysis.json"
HTML_OUT = OUT_DIR / "weekly_review_teams_rich_copy_260810_260816_actual.html"
MD_OUT = OUT_DIR / "weekly_review_teams_copy_260810_260816_actual.md"
TXT_OUT = OUT_DIR / "weekly_review_teams_copy_260810_260816_actual.txt"


def data_uri(img, size=(420, 420)):
    img = ImageOps.exif_transpose(img).convert("RGB")
    img.thumbnail(size, Image.Resampling.LANCZOS)
    canvas = Image.new("RGB", size, "white")
    canvas.paste(img, ((size[0] - img.width) // 2, (size[1] - img.height) // 2))
    buf = BytesIO()
    canvas.save(buf, "JPEG", quality=88, optimize=True)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


def excel_img(code):
    if not TREND.exists():
        return None
    wb = load_workbook(TREND, read_only=False, data_only=True)
    try:
        candidates = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and code in cell.value:
                        rows.append(cell.row)
            if not rows:
                continue
            for img in getattr(ws, "_images", []):
                try:
                    r = img.anchor._from.row + 1
                    c = img.anchor._from.col + 1
                    dist = min(abs(r - hit) for hit in rows)
                    if dist <= 16:
                        candidates.append((dist, abs(c - 7), img))
                except Exception:
                    pass
        if not candidates:
            return None
        candidates.sort(key=lambda x: (x[0], x[1]))
        return data_uri(Image.open(BytesIO(candidates[0][2]._data())))
    finally:
        wb.close()


def musinsa_img(product_id):
    page = urlopen(Request(f"https://www.musinsa.com/products/{product_id}", headers={"User-Agent": "Mozilla/5.0"}), timeout=25).read().decode("utf-8", "ignore")
    m = re.search(r'property=["\']og:image["\']\s+content=["\']([^"\']+)', page) or re.search(r'content=["\']([^"\']+)["\']\s+property=["\']og:image["\']', page)
    if not m:
        return None
    url = html.unescape(m.group(1))
    if url.startswith("//"):
        url = "https:" + url
    raw = urlopen(Request(url, headers={"User-Agent": "Mozilla/5.0"}), timeout=25).read()
    return data_uri(Image.open(BytesIO(raw)))


def direct_img(url):
    raw = urlopen(Request(url, headers={"User-Agent": "Mozilla/5.0"}), timeout=25).read()
    return data_uri(Image.open(BytesIO(raw)))


def placeholder(code):
    img = Image.new("RGB", (420, 260), "#f8f6f0")
    d = ImageDraw.Draw(img)
    try:
        big = ImageFont.truetype("arial.ttf", 32)
        small = ImageFont.truetype("arial.ttf", 18)
    except Exception:
        big = ImageFont.load_default()
        small = ImageFont.load_default()
    d.rectangle((0, 0, 419, 259), outline="#d8d1c5", width=2)
    d.text((24, 70), code, fill="#222", font=big)
    d.text((24, 128), "이미지 확인 필요", fill="#8f3f2f", font=small)
    return data_uri(img, (420, 260))


def img_for(code):
    ids = {
        "WA2602STC3": "6880006",
        "WA2603CD51": "7084458",
    }
    external = {
        "WA2603ST16": "https://img3.momoshop.com.tw/goodsimg/0015/099/645/spec/15099645_01_002_R.jpg?t=1774865661",
    }
    src = excel_img(code)
    if not src and code in ids:
        try:
            src = musinsa_img(ids[code])
        except Exception:
            src = None
    if not src and code in external:
        try:
            src = direct_img(external[code])
        except Exception:
            src = None
    return src or placeholder(code)


IMAGES = {code: img_for(code) for code in ["WA2602STC3", "WA2602STE3", "WA2603CD51", "WA2603ST16", "WA2601LT15", "WA2601CRC1"]}


plain = """■ 8월2주차 어패럴 주간 판매 요약 (26SS + 26FW)
기간: 2026.08.10~08.16 (ERP 기준일 2026.08.17 / 전주 2026.08.03~08.09)
전체 주간판매 6.36억 (26SS+26FW, ACC 포함 / 전주 7.75억 대비 -17.9%)
└ 26SS 5.83억 = APP 4.99억 + ACC 0.84억
└ 26FW 0.53억 = APP 0.38억 + ACC 0.15억

1. 유니섹스/우먼스 전체 전주 판매현황
└ 26SS APP 4.99억 (전년비 -3.3% / 전주대비 -21.0%)
   유니 3.05억(전년비 -18.0%/전주대비 -19.8%) : 우먼 1.88억(전년비 +40.7%/전주대비 -23.4%) = 61.8 : 38.2
└ 26FW APP 0.38억 (전년비 -67.2% / 전주대비 +18.8%)
   유니 0.16억(전년비 -77.8%/전주대비 +28.7%) : 우먼 0.22억(전년비 -50.4%/전주대비 +12.7%) = 41.6 : 58.4
└ 이번 주는 26SS 성숙 시즌 매출 하락이 전체를 끌어내렸고, 26FW는 입고 초기라 절대 금액은 작지만 신규 입고 품번 중심으로 전주대비 플러스 전환.

2. 복종별 판매 현황
└ 26SS 유니 ST 1.81억/4,211pcs, WoW -24.2% — 3PACK/베이직 라인 하락이 크고, 먼작귀 콜라보만 역으로 증가.
└ 26SS 우먼 ST 0.85억/2,062pcs, WoW -26.7% — 상위 반팔 다수가 동반 둔화, ST79는 88장→9장으로 급락.
└ 26SS SO는 유니 0.43억(-35.8%), 우먼 0.19억(-42.2%)로 동시 하락 — 숏팬츠/버뮤다류 시즌 피크아웃 영향 확인 필요.
└ 26SS 유니 LT 0.14억(+93.6%), CR 0.09억(+110.5%)은 반등했지만 절대 볼륨은 제한적.
└ 26FW 우먼 CD는 신규 런칭 96pcs/970만원으로 이번 주 FW 증가의 핵심. 반면 26FW 우먼 ST는 0.10억, WoW -47.7%로 기존 선입고분 조정이 큼.

3. 복종 내 TOP 아이템/판매 채널
※ 26SS 유니 ST — 먼작귀 콜라보 증가 1위
WA2602STC3 먼작귀 친구들 그래픽 반팔티셔츠 105장→211장, +562만원(+94.8%)
└ 판매채널: 직영점 25%, 온라인(자사몰) 17%, 백화점 16%로 분산 반응.

※ 26SS 유니 ST — 하락 영향 1위
WA2602STE3 3PACK 반팔 티셔츠 483장→199장, -1,549만원(-63.0%)
└ 판매채널: 직영점 64%, 면세점 26% 비중. 직영/면세 소진 속도 둔화 확인 필요.

※ 26FW 우먼 CD — 신규 런칭 최다 판매
WA2603CD51 우먼스 릴리와펜 라운드넥 가디건 신규, 이번 주 45장/428만원
└ 판매채널: 백화점 55%, 직영점 29%, 아울렛 7%로 오프라인 초반 반응 우위.

※ 26FW 유니 ST — 신규 입고 반응
WA2603ST16 멀티 그래픽 반팔 티셔츠 신규, 이번 주 58장/261만원
└ 판매채널: 직영점 44%, 백화점 32%, 아울렛 17% 중심.

※ 26SS 유니 LT/CR — 온라인·콜라보 반등
WA2601LT15 원포인트 스트라이프 롱슬리브 3장→67장, +422만원(+1,995.1%)
└ 판매채널: 온라인(무신사) 55%, 해외 위탁 26%, 온라인(위탁몰) 12%.
WA2601CRC1 먼작귀 친구들 그래픽 맨투맨 18장→54장, +325만원(+185.4%)
└ 판매채널: 백화점 26%, 온라인(위탁몰) 25%, 직영점 18%.

4. 액션/확인사항
└ 26SS ST는 STE3/STE1 등 팩 티셔츠와 일반 베이직 라인의 하락 원인을 직영·면세 재고/노출 기준으로 분리 확인 필요.
└ 먼작귀 콜라보는 STC3, STC2, STC1 모두 증가 흐름이어서 오프라인 확대 이후 추가 추적 필요.
└ 26FW 우먼 CD는 백화점 초반 반응이 가장 뚜렷해 CD51/CD55/CD53의 점별 재고 배분과 추가 노출 우선 검토.
└ SO류는 유니/우먼 모두 낙폭이 커서 주차성 시즌 피크아웃인지, 채널별 할인/노출 축소인지 다음 주까지 확인.

작성: 상품기획팀 / ERP 기준 2026-08-17 · 전년비는 주간 리뷰 파일의 작년 비교 시트 기준 · 사진=본문에 언급된 주요 품번만 표시"""


def esc_lines(text):
    out = []
    for line in text.splitlines():
        if not line:
            out.append("<p>&nbsp;</p>")
        elif line.startswith("■"):
            out.append(f"<p class='title'>{html.escape(line)}</p>")
        elif line[0:2] in ("1.", "2.", "3.", "4."):
            out.append(f"<p class='section'>{html.escape(line)}</p>")
        elif line.startswith("※"):
            out.append(f"<p class='issue'>{html.escape(line)}</p>")
        else:
            out.append(f"<p>{html.escape(line)}</p>")
        for code in IMAGES:
            if line.startswith(code):
                out.append(f"<p><img class='product' src='{IMAGES[code]}' alt='{code}'></p>")
    return "\n".join(out)


body = esc_lines(plain)
html_doc = f"""<!doctype html>
<html lang="ko"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>8월2주차 어패럴 주간 판매 요약</title>
<style>
body{{margin:0;background:#eeeae3;font-family:'Malgun Gothic',Arial,sans-serif;color:#222;line-height:1.55}}
.page{{max-width:900px;margin:0 auto;padding:28px 18px 48px}}
.toolbar{{display:flex;justify-content:space-between;gap:12px;margin-bottom:12px;color:#5b554d;font-size:13px}}
button{{border:1px solid #222;background:#222;color:#fff;border-radius:6px;padding:9px 14px;font-weight:700;cursor:pointer}}
#copyArea{{background:#fff;border:1px solid #ddd6cc;padding:30px;box-shadow:0 8px 24px rgba(40,36,30,.08)}}
p{{margin:0 0 10px;white-space:pre-wrap}}
.title{{font-weight:800;font-size:22px;margin-bottom:16px}}
.section{{font-weight:800;font-size:18px;margin-top:20px;padding-top:12px;border-top:1px solid #e7e0d6}}
.issue{{font-weight:800;margin-top:16px}}
.product{{display:block;width:210px;max-width:70%;height:auto;margin:6px 0 16px 20px;border:1px solid #ded8cf;border-radius:4px;background:#fff}}
@media(max-width:640px){{#copyArea{{padding:20px}}.product{{margin-left:0;width:200px}}}}
</style></head><body><div class="page">
<div class="toolbar"><span>흰 영역을 복사하거나 버튼으로 Teams에 붙여넣기</span><button id="copyBtn">리치 복사</button></div>
<div id="copyArea" contenteditable="true">{body}</div>
</div><script>
document.getElementById('copyBtn').onclick=async()=>{{
 const area=document.getElementById('copyArea');
 try{{
  await navigator.clipboard.write([new ClipboardItem({{'text/html':new Blob([area.innerHTML],{{type:'text/html'}}),'text/plain':new Blob([area.innerText],{{type:'text/plain'}})}})]);
 }}catch(e){{const r=document.createRange();r.selectNodeContents(area);const s=window.getSelection();s.removeAllRanges();s.addRange(r);document.execCommand('copy');}}
 document.getElementById('copyBtn').textContent='복사 완료';
}};
</script></body></html>"""

md = plain
for code, src in IMAGES.items():
    md = md.replace(f"\n{code} ", f"\n![{code}]({src})\n\n{code} ")

HTML_OUT.write_text(html_doc, encoding="utf-8")
MD_OUT.write_text(md, encoding="utf-8")
TXT_OUT.write_text(plain, encoding="utf-8")

print(HTML_OUT)
print(MD_OUT)
print(TXT_OUT)
for code, src in IMAGES.items():
    print(code, len(src))
