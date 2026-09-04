import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

BASE = Path("\ucd94\uac00 \ub370\uc774\ud130") / "\ud310\ub9e4 \ub370\uc774\ud130"
CUR_DIR = BASE / "260810~260816"
PREV_DIR = BASE / "260803~260809"
CUR_RAW = CUR_DIR / "\ud310\ub9e4\uc9d1\uacc4\ud604\ud669 260810~260816.xlsx"
PREV_RAW = PREV_DIR / "\ud310\ub9e4\uc9d1\uacc4\ud604\ud669 260803~260809.xlsx"
REVIEW = CUR_DIR / "\u2605\uc640\ud0a4\uc70c\ub9ac_26FW \uc8fc\uac04 \ud310\ub9e4\ub9ac\ubdf0_8\uc6d42\uc8fc\ucc28_NEW.xlsx"

APP = "\uc758\ub958"
ACC = "ACC/\uc6a9\ud488"
UNISEX = "\uc720\ub2c8\uc139\uc2a4"
WOMENS = "\uc6b0\uba3c\uc2a4"
COMMON = "\uacf5\ud1b5"
FEMALE = "\uc5ec\uc131"

APP_CATS = set("JK VT JP CT CR HD HZ KN CD KT LT SH BL ST SS OP SK PT DP SO".split())


def num(x):
    try:
        if x is None or x == "":
            return 0.0
        return float(str(x).replace(",", ""))
    except Exception:
        return 0.0


def season_from_code(code):
    if not isinstance(code, str):
        return None
    if code.startswith("WA2603"):
        return "26FW"
    if code.startswith("WA2601") or code.startswith("WA2602"):
        return "26SS"
    return None


def cat_from_code(code):
    return code[6:8] if isinstance(code, str) and len(code) >= 8 else ""


def canonical_gender(g):
    if g in (UNISEX, COMMON):
        return UNISEX
    if g in (WOMENS, FEMALE):
        return WOMENS
    return g or "\ubbf8\ubd84\ub958"


def build_style_map():
    style_map = {}
    wb = load_workbook(REVIEW, read_only=True, data_only=True)
    ws = wb["\ud488\ubc88\uc885\ud569\uc9d1\uacc4\ud45c(\uae08\ub144)"]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 3:
            continue
        code = row[16] if len(row) > 16 else None
        if not isinstance(code, str) or not code.startswith("WA26"):
            continue
        code = code.strip()
        cat = row[5] or cat_from_code(code)
        style_map.setdefault(
            code,
            {
                "season": row[14] or season_from_code(code),
                "gender": canonical_gender(row[2] or row[15]),
                "appacc": row[3] or (APP if cat in APP_CATS else ACC),
                "cat": cat,
            },
        )
    wb.close()
    return style_map


STYLE_MAP = build_style_map()


def meta(code):
    m = STYLE_MAP.get(code, {})
    cat = m.get("cat") or cat_from_code(code)
    return {
        "season": m.get("season") or season_from_code(code),
        "gender": canonical_gender(m.get("gender")),
        "appacc": m.get("appacc") or (APP if cat in APP_CATS else ACC),
        "cat": cat,
    }


def read_raw(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 2:
            continue
        code = row[4]
        if not isinstance(code, str) or not code.startswith("WA26"):
            continue
        se = season_from_code(code)
        if se not in ("26SS", "26FW"):
            continue
        qty = num(row[10])
        amt = num(row[11])
        if qty == 0 and amt == 0:
            continue
        mm = meta(code)
        if mm["season"] not in ("26SS", "26FW"):
            mm["season"] = se
        rows.append(
            {
                "store_type": row[1] or "\ubbf8\ubd84\ub958",
                "store": row[3] or "",
                "code": code,
                "name": row[5] or "",
                "qty": qty,
                "amt": amt,
                **mm,
            }
        )
    wb.close()
    return rows


def agg(rows, keys):
    out = defaultdict(lambda: {"qty": 0.0, "amt": 0.0})
    for r in rows:
        out[tuple(r[k] for k in keys)]["qty"] += r["qty"]
        out[tuple(r[k] for k in keys)]["amt"] += r["amt"]
    return out


def pct(c, p):
    return None if not p else (c - p) / p * 100


def as_row(cur, prev):
    return {
        "cur_qty": round(cur.get("qty", 0)),
        "cur_amt": round(cur.get("amt", 0)),
        "prev_qty": round(prev.get("qty", 0)),
        "prev_amt": round(prev.get("amt", 0)),
        "diff_qty": round(cur.get("qty", 0) - prev.get("qty", 0)),
        "diff_amt": round(cur.get("amt", 0) - prev.get("amt", 0)),
        "wow": pct(cur.get("amt", 0), prev.get("amt", 0)),
    }


def channel_top(rows, code):
    d = agg([r for r in rows if r["code"] == code], ["store_type"])
    total = sum(v["amt"] for v in d.values())
    tops = []
    for (ch,), v in sorted(d.items(), key=lambda kv: kv[1]["amt"], reverse=True)[:3]:
        tops.append({"channel": ch, "qty": round(v["qty"]), "amt": round(v["amt"]), "share": (v["amt"] / total * 100 if total else 0)})
    return tops


def main():
    cur = read_raw(CUR_RAW)
    prev = read_raw(PREV_RAW)

    totals_cur = agg(cur, ["season", "appacc"])
    totals_prev = agg(prev, ["season", "appacc"])
    gender_cur = agg([r for r in cur if r["appacc"] == APP], ["season", "gender"])
    gender_prev = agg([r for r in prev if r["appacc"] == APP], ["season", "gender"])
    cat_cur = agg([r for r in cur if r["appacc"] == APP], ["season", "gender", "cat"])
    cat_prev = agg([r for r in prev if r["appacc"] == APP], ["season", "gender", "cat"])
    style_cur = agg([r for r in cur if r["appacc"] == APP], ["season", "gender", "cat", "code", "name"])
    style_prev = agg([r for r in prev if r["appacc"] == APP], ["season", "gender", "cat", "code", "name"])

    totals = {}
    for se in ("26SS", "26FW"):
        for aa in (APP, ACC):
            totals[f"{se}_{aa}"] = as_row(totals_cur.get((se, aa), {}), totals_prev.get((se, aa), {}))

    genders = {}
    for se in ("26SS", "26FW"):
        for g in (UNISEX, WOMENS):
            genders[f"{se}_{g}"] = as_row(gender_cur.get((se, g), {}), gender_prev.get((se, g), {}))

    cats = []
    for k in set(cat_cur) | set(cat_prev):
        row = as_row(cat_cur.get(k, {}), cat_prev.get(k, {}))
        row["season"], row["gender"], row["cat"] = k[0], k[1], k[2]
        cats.append(row)
    cats.sort(key=lambda x: abs(x["diff_amt"]), reverse=True)

    styles = []
    for k in set(style_cur) | set(style_prev):
        row = as_row(style_cur.get(k, {}), style_prev.get(k, {}))
        row.update({"season": k[0], "gender": k[1], "cat": k[2], "code": k[3], "name": k[4]})
        styles.append(row)
    styles_best = sorted(styles, key=lambda x: x["cur_amt"], reverse=True)
    styles_up = sorted(styles, key=lambda x: x["diff_amt"], reverse=True)
    styles_down = sorted(styles, key=lambda x: x["diff_amt"])

    selected_codes = []
    for seq in (styles_best[:10], styles_up[:10], styles_down[:10]):
        for s in seq:
            if s["code"] not in selected_codes:
                selected_codes.append(s["code"])

    channels = {code: channel_top(cur, code) for code in selected_codes}

    out = {
        "rows": {"current": len(cur), "previous": len(prev), "style_map": len(STYLE_MAP)},
        "totals": totals,
        "genders": genders,
        "cats": cats[:30],
        "styles_best": styles_best[:30],
        "styles_up": styles_up[:30],
        "styles_down": styles_down[:30],
        "channels": channels,
    }
    Path("tmp").mkdir(exist_ok=True)
    (Path("tmp") / "weekly_actual_analysis.json").write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(out, ensure_ascii=True, indent=2)[:12000])


if __name__ == "__main__":
    main()
