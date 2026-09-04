import base64
import html
import re
from io import BytesIO
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont, ImageOps


ROOT = Path.cwd()
BASE = ROOT / "추가 데이터"
OUT_DIR = BASE / "판매 데이터" / "260810~260816"
OUT = OUT_DIR / "weekly_review_teams_rich_copy_260810_260816_inline.html"
OUT_MAIN = OUT_DIR / "weekly_review_teams_rich_copy_260810_260816.html"


def image_to_data_uri(img: Image.Image, max_size=(520, 520)) -> str:
    img = ImageOps.exif_transpose(img).convert("RGB")
    img.thumbnail(max_size, Image.Resampling.LANCZOS)
    canvas = Image.new("RGB", (max_size[0], max_size[1]), "white")
    canvas.paste(img, ((max_size[0] - img.width) // 2, (max_size[1] - img.height) // 2))
    buf = BytesIO()
    canvas.save(buf, "JPEG", quality=88, optimize=True)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


def extract_nearest_excel_image(xlsx: Path, code: str) -> str | None:
    wb = load_workbook(xlsx, read_only=False, data_only=True)
    try:
        candidates = []
        for ws in wb.worksheets:
            hit_rows = []
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and code in cell.value:
                        hit_rows.append(cell.row)
            if not hit_rows:
                continue
            for img in getattr(ws, "_images", []):
                try:
                    img_row = img.anchor._from.row + 1
                    img_col = img.anchor._from.col + 1
                    dist = min(abs(img_row - r) for r in hit_rows)
                    if dist <= 14:
                        candidates.append((dist, abs(img_col - 7), img))
                except Exception:
                    continue
        if not candidates:
            return None
        candidates.sort(key=lambda x: (x[0], x[1]))
        raw = candidates[0][2]._data()
        return image_to_data_uri(Image.open(BytesIO(raw)))
    finally:
        wb.close()


def musinsa_product_image(product_id: str) -> str | None:
    url = f"https://www.musinsa.com/products/{product_id}"
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    page = urlopen(req, timeout=25).read().decode("utf-8", "ignore")
    m = re.search(r'property=["\']og:image["\']\s+content=["\']([^"\']+)', page)
    if not m:
        m = re.search(r'content=["\']([^"\']+)["\']\s+property=["\']og:image["\']', page)
    if not m:
        return None
    img_url = html.unescape(m.group(1))
    if img_url.startswith("//"):
        img_url = "https:" + img_url
    img_req = Request(img_url, headers={"User-Agent": "Mozilla/5.0"})
    raw = urlopen(img_req, timeout=25).read()
    return image_to_data_uri(Image.open(BytesIO(raw)))


def direct_image(url: str) -> str | None:
    req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
    raw = urlopen(req, timeout=25).read()
    return image_to_data_uri(Image.open(BytesIO(raw)))


def placeholder(code: str, note: str) -> str:
    w, h = 520, 260
    img = Image.new("RGB", (w, h), "#f7f7f4")
    d = ImageDraw.Draw(img)
    try:
        font_big = ImageFont.truetype("arial.ttf", 34)
        font_small = ImageFont.truetype("arial.ttf", 18)
    except Exception:
        font_big = ImageFont.load_default()
        font_small = ImageFont.load_default()
    d.rectangle((0, 0, w - 1, h - 1), outline="#d6d2ca", width=2)
    d.text((28, 64), code, fill="#222222", font=font_big)
    d.text((28, 122), note, fill="#666666", font=font_small)
    d.text((28, 158), "이미지 확인 필요", fill="#9a3f2f", font=font_small)
    return image_to_data_uri(img, max_size=(520, 260))


trend = OUT_DIR / "와키윌리_26SS 전상품 판매추이_260816.xlsx"
images = {
    "WA2602STC3": extract_nearest_excel_image(trend, "WA2602STC3"),
    "WA2602ST79": extract_nearest_excel_image(trend, "WA2602ST79"),
    "WA2603CD51": None,
    "WA2603ST16": None,
}

for code, pid in {"WA2602STC3": "6880006", "WA2602ST79": "6140637", "WA2603CD51": "7084458"}.items():
    if images.get(code):
        continue
    try:
        images[code] = musinsa_product_image(pid)
    except Exception:
        pass

if not images["WA2603ST16"]:
    try:
        images["WA2603ST16"] = direct_image("https://img3.momoshop.com.tw/goodsimg/0015/099/645/spec/15099645_01_002_R.jpg?t=1774865661")
    except Exception:
        pass

if not images["WA2603ST16"]:
    images["WA2603ST16"] = placeholder("WA2603ST16", "26FW 신규 입고 ST")
if not images["WA2603CD51"]:
    images["WA2603CD51"] = placeholder("WA2603CD51", "26FW 신규 런칭 CD")


def product_block(code: str, caption: str) -> str:
    src = images.get(code) or placeholder(code, "이미지 확인 필요")
    return f"""
      <div class="product-card" contenteditable="false">
        <img src="{src}" alt="{html.escape(code)}">
        <div class="product-meta">
          <strong>{html.escape(code)}</strong>
          <span>{html.escape(caption)}</span>
        </div>
      </div>
    """


body = f"""
<div class="report-copy">
  <p class="title">■ 8월2주차 어패럴 주간 판매 요약 (26SS + 26FW)</p>
  <p>기간: 2025.08.18 VS 2026.08.17 (ERP 기준일 2026-08-17)</p>
  <p>전체 주간판매 10.64억 (ACC 포함)<br>
  └ 26SS 5.80억 = APP 4.97억 + ACC 0.83억<br>
  └ 26FW 4.84억 = APP 3.81억 + ACC 1.03억</p>

  <p class="section">▶ 26SS — 성숙 시즌 (누계 판매율 50.1%)</p>
  <p>APP 4.97억 (전년비 -3.3% / 전주대비 -25.2%)<br>
  └ 유니 3.04억(전년비 -18.0%/전주대비 -23.9%) : 우먼 1.87억(전년비 +40.7%/전주대비 -27.3%) = 61.9 : 38.1</p>

  <p class="sub">핵심 포인트</p>
  <p>└ 유니 ST(유니 매출의 59%) WoW -29.9% — 먼작귀 콜라보 라인(STC1~3)은 오히려 증가세, 3PACK/베이직 라인이 크게 빠지며 전체 순감<br>
  └ 우먼 ST(우먼 매출의 45%) WoW -29.5% — WA2602ST79가 88장→9장(-90%)으로 급락, 상위권 다수 동반 하락<br>
  └ 우먼 CD 전년비 +265.7%로 우먼 최고 호조 아이템 (누계 판매율 67.4%, 우먼 평균 62.7% 상회)</p>

  <p class="issue">※ 유니 ST — 먼작귀 콜라보 라인 반응 (전주대비 증가)<br>
  WA2602STC3 105장→209장, +551만원(+93.0%)</p>
  {product_block("WA2602STC3", "먼작귀 콜라보 라인 / 전주대비 증가")}

  <p class="issue">※ 우먼 ST — 전주대비 판매 급락 1위<br>
  WA2602ST79 88장→9장, -346만원(-90.1%)</p>
  {product_block("WA2602ST79", "우먼 ST / 전주대비 급락")}

  <p class="section">▶ 26FW — 입고 초기 시즌 (누계 판매율 1.7%)</p>
  <p>APP 3.81억 (전년비 -67.2% / 전주대비 -5.2%)<br>
  └ (추정) 전년비 낙폭이 큰 것은 순차 입고가 막 시작된 시즌 극초반 특성 — 절대 실적보다 트렌드 위주로 볼 것<br>
  └ 유니 1.58억(전년비 -77.8%/전주대비 -47.2%) : 우먼 2.23억(전년비 -50.4%/전주대비 +129.9%) = 41.5 : 58.5</p>

  <p class="sub">핵심 포인트</p>
  <p>└ 유니는 사실상 ST 단일 아이템만 판매 실적 발생(328장) — 신규 입고분(ST16/ST15)이 증가 견인<br>
  └ 우먼 CD(가디건) 신규 런칭 96장이 우먼 WoW +129.9% 견인 — 전주 판매 0에서 시작한 기저효과 감안 필요<br>
  └ 반면 우먼 ST 4개 품번은 전량 전주대비 하락 — 신규 유입과 기존 아이템 조정이 동시 진행 중</p>

  <p class="issue">※ 유니 ST — 신규 입고 반응<br>
  WA2603ST16 신규 입고, 이번주 58장/261만원</p>
  {product_block("WA2603ST16", "26FW 유니 ST / 신규 입고")}

  <p class="issue">※ 우먼 CD — 신규 런칭 최다 판매<br>
  WA2603CD51 신규 입고, 이번주 45장/428만원</p>
  {product_block("WA2603CD51", "26FW 우먼 CD / 신규 런칭 최다 판매")}

  <p class="section">참고 — ACC</p>
  <p>26SS 0.83억(전년비 -50.4%/전주대비 -17.8%) · 26FW 1.03억(전년비 -67.8%/전주대비 -42.4%)</p>

  <p class="foot">작성: 상품기획팀 / ERP 기준 2026-08-17 (8월2주차, 26SS+26FW 통합) · 사진=전주대비 판매 증가 또는 이슈가 확인된 개별 품번만 표시 · 이미지 출처: 로컬 판매추이 엑셀, 무신사 및 외부 검색 후보</p>
</div>
"""

doc = f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>8월2주차 어패럴 주간 판매 요약</title>
  <style>
    body {{
      margin: 0;
      background: #eeeae3;
      color: #222;
      font-family: "Pretendard", "Malgun Gothic", Arial, sans-serif;
      line-height: 1.55;
    }}
    .page {{
      max-width: 900px;
      margin: 0 auto;
      padding: 28px 18px 48px;
    }}
    .toolbar {{
      display: flex;
      gap: 8px;
      align-items: center;
      justify-content: space-between;
      margin-bottom: 14px;
      color: #5f5a52;
      font-size: 13px;
    }}
    button {{
      border: 1px solid #222;
      background: #222;
      color: #fff;
      border-radius: 6px;
      padding: 9px 14px;
      font-weight: 700;
      cursor: pointer;
    }}
    .report-copy {{
      background: #fff;
      padding: 30px;
      border: 1px solid #ddd6cc;
      box-shadow: 0 8px 24px rgba(40, 36, 30, .08);
      font-size: 16px;
    }}
    p {{ margin: 0 0 14px; }}
    .title {{
      font-size: 22px;
      font-weight: 800;
      margin-bottom: 18px;
    }}
    .section {{
      font-size: 18px;
      font-weight: 800;
      margin-top: 22px;
      padding-top: 14px;
      border-top: 1px solid #e8e2d8;
    }}
    .sub {{
      font-weight: 800;
      margin-bottom: 6px;
    }}
    .issue {{
      font-weight: 700;
      margin-top: 18px;
      margin-bottom: 8px;
    }}
    .product-card {{
      width: 240px;
      margin: 8px 0 18px 22px;
      border: 1px solid #ded8cf;
      background: #fbfaf8;
      border-radius: 6px;
      overflow: hidden;
    }}
    .product-card img {{
      display: block;
      width: 100%;
      height: auto;
      background: #fff;
    }}
    .product-meta {{
      padding: 8px 10px 10px;
      font-size: 12px;
      color: #555;
    }}
    .product-meta strong {{
      display: block;
      color: #222;
      font-size: 13px;
    }}
    .foot {{
      margin-top: 22px;
      color: #666;
      font-size: 13px;
    }}
    @media (max-width: 640px) {{
      .report-copy {{ padding: 20px; font-size: 15px; }}
      .product-card {{ width: 220px; margin-left: 0; }}
    }}
  </style>
</head>
<body>
  <div class="page">
    <div class="toolbar">
      <span>아래 흰 영역을 드래그 복사하거나 버튼으로 복사해서 Teams에 붙여넣기</span>
      <button id="copyBtn">리치 복사</button>
    </div>
    <div id="copyArea" contenteditable="true">
      {body}
    </div>
  </div>
  <script>
    document.getElementById('copyBtn').addEventListener('click', async () => {{
      const area = document.getElementById('copyArea');
      const html = area.innerHTML;
      const text = area.innerText;
      try {{
        await navigator.clipboard.write([
          new ClipboardItem({{
            'text/html': new Blob([html], {{ type: 'text/html' }}),
            'text/plain': new Blob([text], {{ type: 'text/plain' }})
          }})
        ]);
        document.getElementById('copyBtn').textContent = '복사 완료';
      }} catch (e) {{
        const range = document.createRange();
        range.selectNodeContents(area);
        const sel = window.getSelection();
        sel.removeAllRanges();
        sel.addRange(range);
        document.execCommand('copy');
        document.getElementById('copyBtn').textContent = '복사 완료';
      }}
    }});
  </script>
</body>
</html>
"""

OUT.write_text(doc, encoding="utf-8")
OUT_MAIN.write_text(doc, encoding="utf-8")

print(OUT)
print(OUT_MAIN)
for code, src in images.items():
    print(code, "ok" if src and "base64" in src else "missing", len(src or ""))
