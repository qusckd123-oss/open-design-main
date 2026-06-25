from __future__ import annotations

import html
import math
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


COMMON = "\uacf5\ud1b5"


@dataclass
class StyleMetrics:
    code: str
    name: str = ""
    bta: str = "미분류"
    broad: str = "기타"
    incoming_qty: float = 0
    incoming_amt: float = 0
    weekly_units: float = 0
    weekly_net: float = 0
    weekly_tag: float = 0
    weekly_current: float = 0
    cum_units: float = 0
    cum_net: float = 0
    sell_through_weight_sum: float = 0
    sell_through_weight: float = 0

    @property
    def asp(self) -> float:
        return self.weekly_net / self.weekly_units if self.weekly_units else 0

    @property
    def discount(self) -> float:
        return 1 - (self.weekly_net / self.weekly_tag) if self.weekly_tag else 0

    @property
    def sell_through(self) -> float:
        if self.sell_through_weight:
            return self.sell_through_weight_sum / self.sell_through_weight
        if self.incoming_qty:
            return self.cum_units / self.incoming_qty * 100
        return 0


def num(v: Any) -> float:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if math.isnan(v):
            return 0
        return float(v)
    return 0


def find_data_dir(root: Path) -> Path:
    add = max([d for d in root.iterdir() if d.is_dir()], key=lambda d: sum(1 for _ in d.glob("*.xlsx")))
    return max([d for d in add.iterdir() if d.is_dir()], key=lambda d: sum(1 for _ in d.glob("*.xlsx")))


def file_by_token(base: Path, *tokens: str, min_size: int = 0, max_size: int | None = None) -> Path:
    matches = [p for p in base.glob("*.xlsx") if all(t in p.name for t in tokens)]
    if min_size:
        matches = [p for p in matches if p.stat().st_size >= min_size]
    if max_size is not None:
        matches = [p for p in matches if p.stat().st_size <= max_size]
    if not matches:
        raise FileNotFoundError(tokens)
    return matches[0]


def is_st(code: Any) -> bool:
    return isinstance(code, str) and len(code) >= 8 and code[6:8] == "ST"


def broad_group(row: dict[str, str]) -> str:
    text = " ".join(str(v or "") for v in row.values())
    if any(k in text for k in ["스몰", "와펜", "로고", "키키"]):
        if "스트라이프" not in text:
            return "스몰로고/와펜"
    if "앞판" in text:
        return "앞판 그래픽"
    if "뒷판" in text:
        return "뒷판 그래픽"
    if any(k in text for k in ["스트라이프", "링거", "단가라"]):
        return "스트라이프/링거"
    if any(k in text for k in ["워싱", "피그먼트", "다잉", "물나염"]):
        return "워싱/다잉/물나염"
    if any(k in text for k in ["쿨", "기능", "소로나", "쿨맥스"]):
        return "기능성/쿨링"
    if any(k in text for k in ["그래픽", "타이포", "릴리", "캐릭터"]):
        return "기타 그래픽/타이포"
    return "기타"


def load_classification(path: Path) -> dict[str, dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    out: dict[str, dict[str, str]] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        header_i = None
        for i, row in enumerate(rows):
            vals = [str(v or "") for v in row]
            if "BTA 구분" in vals and any(v in vals for v in ["품번", "확정품번"]):
                header_i = i
                break
        if header_i is None:
            continue
        headers = [str(v or "") for v in rows[header_i]]
        for row in rows[header_i + 1 :]:
            rec = {h: str(v or "") for h, v in zip(headers, row)}
            code = rec.get("품번") or rec.get("확정품번")
            if not is_st(code):
                continue
            out[code] = {
                "bta": (rec.get("BTA 구분") or "미분류").strip(),
                "name": rec.get("상품명") or "",
                "broad": broad_group(rec),
            }
    return out


def load_planning(path: Path, prefix: str, season: str, classes: dict[str, dict[str, str]]) -> dict[str, StyleMetrics]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    styles: dict[str, StyleMetrics] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = row[1]
        if not (is_st(code) and code.startswith(prefix) and row[15] == season and row[16] == COMMON):
            continue
        order_qty = num(row[19])
        if order_qty <= 0:
            continue
        style = styles.setdefault(code, StyleMetrics(code=code))
        cls = classes.get(code, {})
        style.bta = cls.get("bta", style.bta)
        style.name = cls.get("name", style.name)
        style.broad = cls.get("broad", style.broad)
        style.incoming_qty += order_qty
        style.incoming_amt += num(row[23]) or order_qty * num(row[7])
        style.cum_units += num(row[47])
        style.cum_net += num(row[53])
        rate = num(row[58])
        if rate:
            style.sell_through_weight_sum += rate * order_qty
            style.sell_through_weight += order_qty
        if not style.name:
            style.name = str(row[1] or "")
    return styles


def apply_weekly_sales(path: Path, styles: dict[str, StyleMetrics]) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = row[2]
        if code not in styles:
            continue
        s = styles[code]
        if not s.name:
            s.name = str(row[4] or "")
        units = num(row[7])
        tag = num(row[5]) * units
        current = num(row[8])
        net = num(row[9])
        s.weekly_units += units
        s.weekly_tag += tag
        s.weekly_current += current
        s.weekly_net += net


def aggregate(styles: dict[str, StyleMetrics], key: str | None = None) -> dict[str, StyleMetrics]:
    groups: dict[str, StyleMetrics] = {}
    for s in styles.values():
        k = getattr(s, key) if key else "TOTAL"
        g = groups.setdefault(k, StyleMetrics(code=k))
        for attr in [
            "incoming_qty",
            "incoming_amt",
            "weekly_units",
            "weekly_net",
            "weekly_tag",
            "weekly_current",
            "cum_units",
            "cum_net",
            "sell_through_weight_sum",
            "sell_through_weight",
        ]:
            setattr(g, attr, getattr(g, attr) + getattr(s, attr))
    return groups


def krw_m(v: float) -> str:
    return f"{v / 1_000_000:,.1f}백만"


def qty(v: float) -> str:
    return f"{v:,.0f}"


def pct(v: float) -> str:
    return f"{v:.1f}%"


def metric_row(label: str, y25: float | str, y26: float | str, delta: float | str = "") -> str:
    return f"<tr><th>{label}</th><td>{y25}</td><td>{y26}</td><td>{delta}</td></tr>"


def top_table(title: str, rows: list[StyleMetrics]) -> str:
    body = []
    for s in rows:
        body.append(
            "<tr>"
            f"<td><b>{html.escape(s.code)}</b><br><span>{html.escape(s.name)}</span></td>"
            f"<td>{html.escape(s.bta)}</td><td>{html.escape(s.broad)}</td>"
            f"<td class='num'>{krw_m(s.weekly_net)}</td><td class='num'>{qty(s.weekly_units)}</td>"
            f"<td class='num'>{krw_m(s.cum_net)}</td><td class='num'>{pct(s.sell_through)}</td>"
            f"<td class='num'>{pct(s.discount * 100)}</td><td class='num'>{s.asp:,.0f}</td>"
            f"<td class='num'>{qty(s.incoming_qty)}</td>"
            "</tr>"
        )
    return (
        f"<h3>{title}</h3><table><tr><th>품번/상품명</th><th>B/T/A</th><th>유형</th>"
        "<th>주간 실판매</th><th>주간 수량</th><th>누계 실판매</th><th>누계판매율</th>"
        "<th>주간 할인율</th><th>ASP</th><th>입고</th></tr>"
        + "".join(body)
        + "</table>"
    )


def comparison_table(title: str, g25: dict[str, StyleMetrics], g26: dict[str, StyleMetrics]) -> str:
    keys = sorted(set(g25) | set(g26), key=lambda k: (g26.get(k, StyleMetrics(k)).weekly_net - g25.get(k, StyleMetrics(k)).weekly_net), reverse=True)
    rows = []
    for k in keys:
        a = g25.get(k, StyleMetrics(k))
        b = g26.get(k, StyleMetrics(k))
        delta = b.weekly_net - a.weekly_net
        cls = "up" if delta >= 0 else "down"
        rows.append(
            f"<tr><th>{html.escape(k)}</th><td>{sum(1 for s in styles25.values() if getattr(s, title) == k) if title in ['bta','broad'] else ''}</td>"
            f"<td>{sum(1 for s in styles26.values() if getattr(s, title) == k) if title in ['bta','broad'] else ''}</td>"
            f"<td class='num'>{krw_m(a.weekly_net)}</td><td class='num'>{krw_m(b.weekly_net)}</td>"
            f"<td class='num {cls}'>{krw_m(delta)}</td><td class='num'>{qty(a.weekly_units)}</td>"
            f"<td class='num'>{qty(b.weekly_units)}</td><td class='num'>{pct(a.sell_through)}</td>"
            f"<td class='num'>{pct(b.sell_through)}</td><td class='num'>{pct(a.discount * 100)}</td>"
            f"<td class='num'>{pct(b.discount * 100)}</td></tr>"
        )
    return (
        "<table><tr><th>구분</th><th>25 스타일</th><th>26 스타일</th><th>25 주간 실판매</th>"
        "<th>26 주간 실판매</th><th>증감</th><th>25 수량</th><th>26 수량</th>"
        "<th>25 누계판매율</th><th>26 누계판매율</th><th>25 할인율</th><th>26 할인율</th></tr>"
        + "".join(rows)
        + "</table>"
    )


root = Path.cwd()
base = find_data_dir(root)
class_path = base / ".tmp_25SS_26SS_item_class.xlsx"
classes = load_classification(class_path)

styles25 = load_planning(file_by_token(base, "250519", "250525", min_size=5_000_000), "WA25", "25SS", classes)
styles26 = load_planning(file_by_token(base, "260518", "260524", min_size=5_000_000), "WA26", "26SS", classes)
apply_weekly_sales(file_by_token(base, "250519", "250525", min_size=100_000, max_size=1_000_000), styles25)
apply_weekly_sales(file_by_token(base, "260518", "260524", min_size=100_000, max_size=1_000_000), styles26)

total25 = aggregate(styles25)["TOTAL"]
total26 = aggregate(styles26)["TOTAL"]
bta25, bta26 = aggregate(styles25, "bta"), aggregate(styles26, "bta")
broad25, broad26 = aggregate(styles25, "broad"), aggregate(styles26, "broad")

weekly_delta = total26.weekly_net - total25.weekly_net
unit_delta = total26.weekly_units - total25.weekly_units
asp_delta = total26.asp - total25.asp
top_growth = sorted(set(broad25) | set(broad26), key=lambda k: broad26.get(k, StyleMetrics(k)).weekly_net - broad25.get(k, StyleMetrics(k)).weekly_net, reverse=True)[:3]
top_decline = sorted(set(broad25) | set(broad26), key=lambda k: broad26.get(k, StyleMetrics(k)).weekly_net - broad25.get(k, StyleMetrics(k)).weekly_net)[:3]

summary = (
    f"26SS 유니섹스 ST 주간 실판매액은 {krw_m(total26.weekly_net)}으로 25SS {krw_m(total25.weekly_net)} 대비 "
    f"{krw_m(weekly_delta)} 변화했습니다. 수량은 {qty(total26.weekly_units)}장으로 {qty(unit_delta)}장, "
    f"ASP는 {total26.asp:,.0f}원으로 {asp_delta:,.0f}원 차이입니다. "
    f"같은 주차 기획현황 기준 누계판매율은 25SS {pct(total25.sell_through)}, 26SS {pct(total26.sell_through)}입니다."
)

facts = [
    f"신장 기여 유형: {', '.join(top_growth)}",
    f"역신장 기여 유형: {', '.join(top_decline)}",
    "스타일 수와 B/T/A 비교는 기획현황상 공통 ST이며 발주수량 0인 품번을 제외했습니다.",
    "누계판매율은 상품MAP/전체 시즌 발주가 아니라 각 비교 주차의 기획현황 스냅샷 값을 발주수량 가중 평균으로 비교했습니다.",
]

metric_rows = [
    metric_row("활성 스타일 수", qty(len(styles25)), qty(len(styles26)), qty(len(styles26) - len(styles25))),
    metric_row("입고 수량", qty(total25.incoming_qty), qty(total26.incoming_qty), qty(total26.incoming_qty - total25.incoming_qty)),
    metric_row("입고 금액", krw_m(total25.incoming_amt), krw_m(total26.incoming_amt), krw_m(total26.incoming_amt - total25.incoming_amt)),
    metric_row("누계판매율", pct(total25.sell_through), pct(total26.sell_through), pct(total26.sell_through - total25.sell_through)),
    metric_row("주간 실판매액", krw_m(total25.weekly_net), krw_m(total26.weekly_net), krw_m(weekly_delta)),
    metric_row("주간 판매수량", qty(total25.weekly_units), qty(total26.weekly_units), qty(unit_delta)),
    metric_row("주간 ASP", f"{total25.asp:,.0f}원", f"{total26.asp:,.0f}원", f"{asp_delta:,.0f}원"),
    metric_row("주간 할인율", pct(total25.discount * 100), pct(total26.discount * 100), pct((total26.discount - total25.discount) * 100)),
]

top25 = sorted(styles25.values(), key=lambda s: s.weekly_net, reverse=True)[:15]
top26 = sorted(styles26.values(), key=lambda s: s.weekly_net, reverse=True)[:15]

bta_sections = []
for bta in ["B", "T", "A", "미분류"]:
    r25 = sorted([s for s in styles25.values() if s.bta == bta], key=lambda s: s.weekly_net, reverse=True)[:8]
    r26 = sorted([s for s in styles26.values() if s.bta == bta], key=lambda s: s.weekly_net, reverse=True)[:8]
    if r25 or r26:
        bta_sections.append(f"<section><h2>{bta}군 주요 판매 품번</h2><div class='grid'>{top_table('25SS', r25)}{top_table('26SS', r26)}</div></section>")

html_doc = f"""<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>ST weekly sales review 250519 260518</title>
<style>
body{{font-family:Arial,'Malgun Gothic',sans-serif;margin:0;background:#f5f5f3;color:#1f2428;line-height:1.45}}
main{{max-width:1280px;margin:0 auto;padding:36px 28px 56px}}
h1{{font-size:30px;margin:0 0 8px}} h2{{font-size:20px;margin:34px 0 12px}} h3{{font-size:15px;margin:16px 0 8px}}
.lead{{font-size:18px;font-weight:700;background:#fff;border-left:5px solid #111;padding:18px 20px;margin:18px 0}}
.notes{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:14px 0 24px}}
.note{{background:#fff;border:1px solid #ddd;padding:12px;border-radius:8px;font-size:13px}}
.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px}}
.kpi{{background:#111;color:#fff;padding:16px;border-radius:8px}} .kpi span{{display:block;font-size:12px;color:#ccc}} .kpi b{{font-size:24px}}
table{{border-collapse:collapse;width:100%;background:#fff;margin:8px 0 18px;font-size:12px}}
th,td{{border:1px solid #ddd;padding:8px 9px;vertical-align:top}} th{{background:#efefea;text-align:left}} td.num{{text-align:right;white-space:nowrap}}
.up{{color:#0a7d38;font-weight:700}} .down{{color:#c53232;font-weight:700}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:16px;align-items:start}}
.small{{color:#666;font-size:12px}} section{{margin-top:24px}}
@media(max-width:900px){{.kpis,.notes,.grid{{grid-template-columns:1fr}} main{{padding:22px 14px}}}}
</style>
</head>
<body><main>
<h1>25SS vs 26SS ST 유니섹스 주간 매출 리뷰</h1>
<p class="small">비교 주차: 25SS 2025-05-19~2025-05-25 / 26SS 2026-05-18~2026-05-24. 기준: 주간 실판매액 중심, 공통 ST, 입고 0 제외, B/T/A 업데이트 반영.</p>
<div class="lead">{html.escape(summary)}</div>
<div class="notes">{''.join(f'<div class="note">{html.escape(x)}</div>' for x in facts)}</div>
<div class="kpis">
<div class="kpi"><span>25SS 주간 실판매</span><b>{krw_m(total25.weekly_net)}</b></div>
<div class="kpi"><span>26SS 주간 실판매</span><b>{krw_m(total26.weekly_net)}</b></div>
<div class="kpi"><span>증감</span><b>{krw_m(weekly_delta)}</b></div>
<div class="kpi"><span>26SS 주간 수량 / ASP</span><b>{qty(total26.weekly_units)}장 / {total26.asp:,.0f}원</b></div>
</div>
<section><h2>주간 매출 비교</h2><table><tr><th>지표</th><th>25SS</th><th>26SS</th><th>증감</th></tr>{''.join(metric_rows)}</table></section>
<section><h2>B/T/A 비교</h2>{comparison_table('bta', bta25, bta26)}</section>
<section><h2>유사 상품군 비교</h2>{comparison_table('broad', broad25, broad26)}</section>
<section><h2>전체 상위 판매 품번</h2><div class="grid">{top_table('25SS Top 15', top25)}{top_table('26SS Top 15', top26)}</div></section>
{''.join(bta_sections)}
<section><h2>최종 판단</h2>
<table><tr><th>구분</th><th>판단</th></tr>
<tr><td>주간 매출</td><td>판단의 우선순위는 누계가 아니라 주간 실판매액입니다. 26SS는 같은 주차 누계판매율을 별도 확인했으며, 판매액 차이는 수량·ASP·할인율·유형 믹스 변화로 해석해야 합니다.</td></tr>
<tr><td>B/T/A</td><td>B/T/A는 업데이트된 아이템 구분 파일을 기준으로 재매핑했습니다. 각 군의 스타일 수와 입고 수량이 다른 만큼, 단순 판매율보다 주간 실판매액과 대표 품번의 판매 깊이를 같이 봐야 합니다.</td></tr>
<tr><td>기획 방향</td><td>신장 유형은 주간 판매 실적이 확인된 범위에서 유지하고, 역신장 유형은 스타일 수·입고 깊이·ASP·할인 의존도를 분리해 리오더/축소 여부를 판단하는 것이 적절합니다.</td></tr>
</table></section>
</main></body></html>"""

out = base / "ST_weekly_sales_review_250519_260518_updated.html"
out.write_text("\ufeff" + html_doc, encoding="utf-8")
print(out)
