from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import matplotlib.dates as mdates
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def main() -> None:
    src = Path(os.environ["FILE"])
    out_dir = src.parent
    out_xlsx = out_dir / "WA2503STE1_TC_reorder_report.xlsx"
    out_png = out_dir / "WA2503STE1_TC_reorder_report.png"

    wb_src = load_workbook(src, data_only=True)
    ws_src = wb_src.active

    initial_stock = int(ws_src["I3"].value)
    sold_to_date = int(ws_src["I4"].value)
    stock_to_date = int(ws_src["I8"].value)
    sell_through = ws_src["I5"].value
    review_date = ws_src["B16"].value or datetime(2026, 5, 19)
    first_sale_date = ws_src["B17"].value or datetime(2025, 6, 15)

    dates: list[datetime] = []
    weekly_src: list[float] = []
    for col in range(13, ws_src.max_column + 1):
        date_value = ws_src.cell(26, col).value
        week_value = ws_src.cell(27, col).value
        if isinstance(date_value, datetime) and isinstance(week_value, (int, float)):
            dates.append(date_value)
            weekly_src.append(float(week_value))

    reorder_1_date = datetime(2026, 5, 24)
    reorder_1_qty = 3000
    reorder_2_actual_date = datetime(2026, 6, 30)
    reorder_2_chart_week = datetime(2026, 7, 5)
    reorder_2_qty = 7000
    final_soldout_week = datetime(2026, 10, 4)
    total_supply = initial_stock + reorder_1_qty + reorder_2_qty

    future_indices = [
        idx
        for idx, date_value in enumerate(dates)
        if reorder_1_date <= date_value <= final_soldout_week
    ]
    future_source_sum = sum(weekly_src[idx] for idx in future_indices)
    remaining_to_sell = total_supply - sold_to_date
    scale_factor = remaining_to_sell / future_source_sum if future_source_sum else 1

    rows = []
    for date_value, source_weekly_sales in zip(dates, weekly_src):
        if date_value < reorder_1_date:
            weekly_sales = source_weekly_sales
            status = "실판매/원천 흐름"
        else:
            weekly_sales = source_weekly_sales * scale_factor
            status = "예측 판매(10/4 소진 보정)"
        rows.append(
            {
                "date": date_value,
                "src_weekly": source_weekly_sales,
                "weekly": weekly_sales,
                "status": status,
            }
        )

    historical_sum = sum(row["weekly"] for row in rows if row["date"] < reorder_1_date)
    historical_scale = sold_to_date / historical_sum if historical_sum else 1
    for row in rows:
        if row["date"] < reorder_1_date:
            row["weekly"] *= historical_scale

    cumulative_sales = 0.0
    cumulative_supply = initial_stock
    for row in rows:
        date_value = row["date"]
        inbound = 0
        event = ""
        if date_value == reorder_1_date:
            inbound += reorder_1_qty
            event = "1차 리오더 입고 3,000"
        if date_value == reorder_2_chart_week:
            inbound += reorder_2_qty
            event = "2차 리오더 입고 7,000 (6/30 입고)"

        cumulative_supply += inbound
        cumulative_sales += row["weekly"]
        if date_value == final_soldout_week:
            cumulative_sales = total_supply

        row.update(
            {
                "inbound": inbound,
                "event": event,
                "cum_sales": cumulative_sales,
                "cum_supply": cumulative_supply,
                "ending_stock": max(cumulative_supply - cumulative_sales, 0),
                "sell_through_total": cumulative_sales / total_supply if total_supply else 0,
            }
        )

    for row in rows:
        if row["date"] == final_soldout_week:
            row["ending_stock"] = 0
            row["cum_sales"] = total_supply
            row["sell_through_total"] = 1

    font_path = Path("C:/Windows/Fonts/malgun.ttf")
    if font_path.exists():
        fm.fontManager.addfont(str(font_path))
        plt.rcParams["font.family"] = "Malgun Gothic"
    plt.rcParams["axes.unicode_minus"] = False

    plot_rows = [row for row in rows if row["date"] >= datetime(2026, 3, 1)]
    fig, ax1 = plt.subplots(figsize=(15, 8), dpi=180)
    fig.patch.set_facecolor("#F7F5EF")
    ax1.set_facecolor("#F7F5EF")

    bar_dates = [row["date"] for row in plot_rows]
    bar_vals = [row["weekly"] for row in plot_rows]
    stock_vals = [row["ending_stock"] for row in plot_rows]
    supply_vals = [row["cum_supply"] for row in plot_rows]
    cumulative_vals = [row["cum_sales"] for row in plot_rows]

    ax2 = ax1.twinx()
    ax2.bar(bar_dates, bar_vals, width=4.8, color="#92B6B1", alpha=0.65, label="주차별 판매/예측")
    ax1.plot(bar_dates, stock_vals, color="#C43D36", linewidth=3.2, marker="o", markersize=3.5, label="기말 재고 런웨이")
    ax1.plot(bar_dates, supply_vals, color="#2F5D7C", linewidth=2.4, linestyle="--", label="누적 공급량")
    ax1.plot(bar_dates, cumulative_vals, color="#303030", linewidth=2.4, label="누적 판매량")

    event_labels = [
        (reorder_1_date, "5/24 1차 리오더\n+3,000장", 0.82),
        (reorder_2_chart_week, "6/30 2차 리오더\n+7,000장", 0.72),
        (final_soldout_week, "10/4주차\n전체 소진", 0.16),
    ]
    for event_date, text, y_ratio in event_labels:
        ax1.axvline(event_date, color="#8B5E34", linewidth=1.3, linestyle=":")
        ax1.annotate(
            text,
            xy=(event_date, max(stock_vals) * y_ratio),
            xytext=(8, 0),
            textcoords="offset points",
            ha="left",
            va="center",
            fontsize=10.5,
            color="#3A2A1A",
            bbox={"boxstyle": "round,pad=0.35", "fc": "#FFF8E5", "ec": "#C6A15B", "lw": 0.8},
        )

    fig.suptitle(
        "WA2503STE1 TC 리오더 진행 근거: 1차 3,000장 + 2차 7,000장 반영 시 10/4주차 전체 소진",
        fontsize=16,
        fontweight="bold",
        y=0.975,
    )
    ax1.text(
        0.0,
        1.015,
        f"현재 판매 {sold_to_date:,}장 / 현재 재고 {stock_to_date:,}장 / 초도 {initial_stock:,}장 / 총 운영공급 {total_supply:,}장",
        transform=ax1.transAxes,
        fontsize=11.5,
        color="#333333",
    )
    ax1.set_ylabel("누적/재고 수량", fontsize=11)
    ax2.set_ylabel("주차별 판매 수량", fontsize=11)
    ax1.yaxis.set_major_formatter(lambda x, _pos: f"{int(x):,}")
    ax2.yaxis.set_major_formatter(lambda x, _pos: f"{int(x):,}")
    ax1.xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=mdates.SU, interval=2))
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
    ax1.grid(axis="y", color="#D9D3C6", linewidth=0.8)
    ax1.set_ylim(0, max(max(supply_vals), total_supply) * 1.08)
    ax2.set_ylim(0, max(bar_vals) * 1.45)
    lines, labels = ax1.get_legend_handles_labels()
    bars, bar_labels = ax2.get_legend_handles_labels()
    ax1.legend(lines + bars, labels + bar_labels, loc="upper left", frameon=False, ncol=4, bbox_to_anchor=(0, -0.12))
    plt.tight_layout(rect=[0, 0.05, 1, 0.925])
    fig.savefig(out_png, bbox_inches="tight")
    plt.close(fig)

    wb = Workbook()
    ws = wb.active
    ws.title = "리오더 근거 그래프"
    ws.sheet_view.showGridLines = False

    navy = "1F4E5F"
    white = "FFFFFF"
    gray = "666666"

    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["J"].width = 30

    ws["A1"] = "WA2503STE1 TC 리오더 진행 근거"
    ws["A1"].font = Font(name="Arial", bold=True, size=20, color="1E1E1E")
    ws["A2"] = "목적: 1차 리오더 3,000장 진행 완료, 2차 리오더 7,000장 진행 필요성을 주차별 판매/재고 런웨이로 제시"
    ws["A2"].font = Font(name="Arial", size=11, color=gray)
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    summary = [
        ("현재 누적 판매", sold_to_date, "장"),
        ("현재 재고", stock_to_date, "장"),
        ("현재 판매율", sell_through, ""),
        ("초도 입고", initial_stock, "장"),
        ("1차 리오더", reorder_1_qty, "장 / 5/24 입고"),
        ("2차 리오더", reorder_2_qty, "장 / 6/30 입고"),
        ("총 운영 공급", total_supply, "장"),
        ("소진 예상", final_soldout_week, "10/4주차"),
    ]
    start_row = 4
    for idx, (label, value, unit) in enumerate(summary):
        row = start_row + (idx // 4) * 3
        col = 1 + (idx % 4) * 2
        ws.cell(row, col).value = label
        ws.cell(row, col).font = Font(name="Arial", bold=True, size=10, color=white)
        ws.cell(row, col).fill = PatternFill("solid", fgColor=navy)
        ws.cell(row, col).alignment = Alignment(horizontal="center")
        ws.cell(row + 1, col).value = value
        ws.cell(row + 1, col).font = Font(name="Arial", bold=True, size=13, color="1E1E1E")
        ws.cell(row + 1, col).alignment = Alignment(horizontal="center")
        if isinstance(value, datetime):
            ws.cell(row + 1, col).number_format = "m/d"
        elif label == "현재 판매율":
            ws.cell(row + 1, col).number_format = "0.0%"
        else:
            ws.cell(row + 1, col).number_format = "#,##0"
        ws.cell(row + 1, col + 1).value = unit
        ws.cell(row + 1, col + 1).font = Font(name="Arial", size=10, color=gray)
        ws.cell(row + 1, col + 1).alignment = Alignment(vertical="center")

    table_row = 12
    headers = ["주차", "구분", "원천 주차 판매", "보정 주차 판매", "입고", "누적 공급", "누적 판매", "기말 재고", "누적 판매율", "이벤트"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(table_row, col)
        cell.value = header
        cell.font = Font(name="Arial", bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, table_row + 1):
        ws.cell(row_idx, 1).value = row["date"]
        ws.cell(row_idx, 2).value = row["status"]
        ws.cell(row_idx, 3).value = row["src_weekly"]
        ws.cell(row_idx, 4).value = f"=IF(A{row_idx}<DATE(2026,5,24),C{row_idx}*{historical_scale},C{row_idx}*{scale_factor})"
        ws.cell(row_idx, 5).value = row["inbound"]
        if row_idx == table_row + 1:
            ws.cell(row_idx, 6).value = f"={initial_stock}+E{row_idx}"
            ws.cell(row_idx, 7).value = f"=D{row_idx}"
        else:
            ws.cell(row_idx, 6).value = f"=F{row_idx - 1}+E{row_idx}"
            ws.cell(row_idx, 7).value = f"=G{row_idx - 1}+D{row_idx}"
        if row["date"] == final_soldout_week:
            ws.cell(row_idx, 7).value = f"={total_supply}"
        ws.cell(row_idx, 8).value = f"=MAX(F{row_idx}-G{row_idx},0)"
        ws.cell(row_idx, 9).value = f"=G{row_idx}/{total_supply}"
        ws.cell(row_idx, 10).value = row["event"]
        for col in range(1, 11):
            ws.cell(row_idx, col).font = Font(name="Arial", size=9)
            ws.cell(row_idx, col).fill = PatternFill("solid", fgColor="FBFAF5" if row_idx % 2 else white)
        ws.cell(row_idx, 1).number_format = "m/d"
        for col in [3, 4, 5, 6, 7, 8]:
            ws.cell(row_idx, col).number_format = "#,##0"
        ws.cell(row_idx, 9).number_format = "0.0%"

    last_row = table_row + len(rows)
    thin = Side(style="thin", color="D8D2C3")
    for row in ws.iter_rows(min_row=table_row, max_row=last_row, min_col=1, max_col=10):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="center")

    note_row = last_row + 3
    notes = [
        "데이터 근거",
        f"- 원본 파일: {src.name}",
        "- 상단 요약의 대만 제외 기준 초도/판매/재고를 사용했습니다.",
        "- 2026-05-24 이후 판매는 원본 주차별 판매 흐름을 유지하되, 10/4주차 총 운영 공급 19,259장 소진 기준으로 보정했습니다.",
        "- 2차 리오더 7,000장은 2026-06-30 입고이나 주차 그래프에서는 해당 주 시작점인 2026-07-05에 반영했습니다.",
    ]
    for idx, text in enumerate(notes):
        ws.cell(note_row + idx, 1).value = text
        ws.cell(note_row + idx, 1).font = Font(name="Arial", bold=(idx == 0), size=10, color="333333")
        ws.merge_cells(start_row=note_row + idx, start_column=1, end_row=note_row + idx, end_column=8)

    img = XLImage(str(out_png))
    img.width = 1000
    img.height = 533
    ws.add_image(img, "A37")

    chart = LineChart()
    chart.title = "재고 런웨이 및 누적 판매/공급"
    chart.y_axis.title = "수량"
    chart.x_axis = DateAxis(crossAx=100)
    chart.x_axis.number_format = "m/d"
    chart.x_axis.majorTimeUnit = "days"
    chart.height = 10
    chart.width = 22
    for col in [6, 7, 8]:
        chart.add_data(Reference(ws, min_col=col, min_row=table_row, max_row=last_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=table_row + 1, max_row=last_row))
    ws.add_chart(chart, "L4")

    bar = BarChart()
    bar.title = "주차별 판매/예측"
    bar.y_axis.title = "수량"
    bar.x_axis.title = "주차"
    bar.height = 10
    bar.width = 22
    bar.add_data(Reference(ws, min_col=4, min_row=table_row, max_row=last_row), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=table_row + 1, max_row=last_row))
    ws.add_chart(bar, "L24")

    ws.freeze_panes = "A13"
    ws.auto_filter.ref = f"A{table_row}:J{last_row}"

    source_ws = wb.create_sheet("원본 요약")
    source_ws.sheet_view.showGridLines = False
    source_ws["A1"] = "원본 요약 및 산출 기준"
    source_ws["A1"].font = Font(name="Arial", bold=True, size=16)
    source_rows = [
        ("품번", "WA2503STE1TC"),
        ("연간 운영 흐름", "25SS WA2502STE1 -> 25FW WA2503STE1 -> 26SS WA2602STE1"),
        ("1차 리오더 품번 생성 기준", "26SS WA2602STE1"),
        ("검토일", review_date),
        ("최초판매일", first_sale_date),
        ("초도 입고", initial_stock),
        ("누적 판매", sold_to_date),
        ("잔여 재고", stock_to_date),
        ("판매율", sell_through),
        ("1차 리오더", reorder_1_qty),
        ("1차 입고일", reorder_1_date),
        ("2차 리오더", reorder_2_qty),
        ("2차 입고일", reorder_2_actual_date),
        ("전체 소진 예상 주차", final_soldout_week),
        ("미래 판매 보정 배율", scale_factor),
    ]
    for row_idx, (key, value) in enumerate(source_rows, 3):
        source_ws.cell(row_idx, 1).value = key
        source_ws.cell(row_idx, 2).value = value
        source_ws.cell(row_idx, 1).font = Font(name="Arial", bold=True)
        source_ws.cell(row_idx, 2).font = Font(name="Arial")
        if isinstance(value, datetime):
            source_ws.cell(row_idx, 2).number_format = "yyyy-mm-dd"
        elif isinstance(value, float) and value < 2:
            source_ws.cell(row_idx, 2).number_format = "0.0%"
        elif isinstance(value, (int, float)):
            source_ws.cell(row_idx, 2).number_format = "#,##0.0"
    source_ws.column_dimensions["A"].width = 32
    source_ws.column_dimensions["B"].width = 48

    wb.save(out_xlsx)
    print(out_xlsx)
    print(out_png)
    print(f"scale_factor={scale_factor:.6f}")
    print(f"historical_scale={historical_scale:.6f}")
    print(f"total_supply={total_supply}")


if __name__ == "__main__":
    main()
