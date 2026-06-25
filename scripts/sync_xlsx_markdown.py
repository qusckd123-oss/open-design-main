from __future__ import annotations

import argparse
import os
import tempfile
import time
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required. Install it with: python -m pip install openpyxl") from exc


DEFAULT_XLSX = Path("추가 데이터") / "(유니) 26SS 와키윌리 리오더 점검 파일 (2부).xlsx"
DEFAULT_MD = Path("추가 데이터") / "(유니) 26SS 와키윌리 리오더 점검 파일 (2부).md"


def format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S") if value.time() else value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value)
    return text.replace("\r\n", "<br>").replace("\n", "<br>").replace("|", "\\|")


def non_empty_bounds(rows: list[list[object]]) -> tuple[int, int]:
    last_row = 0
    last_col = 0
    for row_index, row in enumerate(rows, start=1):
        row_has_value = False
        for col_index, value in enumerate(row, start=1):
            if value is not None and str(value) != "":
                row_has_value = True
                last_col = max(last_col, col_index)
        if row_has_value:
            last_row = row_index
    return last_row, last_col


def markdown_table(rows: Iterable[list[object]], width: int) -> list[str]:
    lines: list[str] = []
    for row_index, row in enumerate(rows):
        cells = [format_cell(row[col_index]) if col_index < len(row) else "" for col_index in range(width)]
        lines.append(f"| {' | '.join(cells)} |")
        if row_index == 0:
            lines.append(f"| {' | '.join(['---'] * width)} |")
    return lines


def convert_xlsx_to_markdown(source: Path, target: Path) -> None:
    source = source.resolve()
    target = target.resolve()
    if not source.exists():
        raise FileNotFoundError(source)

    workbook = load_workbook(source, read_only=True, data_only=True)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines = [
        f"# {source.stem}",
        "",
        f"- Source: `{source.name}`",
        f"- Generated: {generated_at}",
        "",
        "## Sheets",
        "",
    ]

    for index, sheet_name in enumerate(workbook.sheetnames, start=1):
        lines.append(f"- [{sheet_name}](#sheet-{index})")

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        last_row, last_col = non_empty_bounds(rows)

        lines.extend(["", f"## {sheet_name}", ""])
        if last_row == 0 or last_col == 0:
            lines.append("_No data_")
            continue

        table_rows = (row[:last_col] for row in rows[:last_row])
        lines.extend(markdown_table(table_rows, last_col))

    target.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        "w",
        encoding="utf-8",
        newline="\n",
        delete=False,
        dir=target.parent,
        prefix=f".{target.name}.",
        suffix=".tmp",
    ) as tmp:
        tmp.write("\n".join(lines))
        tmp.write("\n")
        tmp_path = Path(tmp.name)

    os.replace(tmp_path, target)


def wait_for_stable_file(path: Path, checks: int = 3, interval: float = 0.8) -> None:
    last: tuple[int, int] | None = None
    stable_count = 0
    while stable_count < checks:
        stat = path.stat()
        current = (stat.st_size, stat.st_mtime_ns)
        if current == last:
            stable_count += 1
        else:
            stable_count = 0
            last = current
        time.sleep(interval)


def watch(source: Path, target: Path, interval: float) -> None:
    print(f"Watching: {source}")
    print(f"Writing:  {target}")
    last_mtime = 0
    while True:
        try:
            current_mtime = source.stat().st_mtime_ns
            if current_mtime != last_mtime:
                wait_for_stable_file(source)
                convert_xlsx_to_markdown(source, target)
                last_mtime = source.stat().st_mtime_ns
                print(f"[{datetime.now().strftime('%H:%M:%S')}] synced")
        except PermissionError:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] waiting for Excel to release the file")
        except KeyboardInterrupt:
            print("Stopped.")
            return
        except Exception as exc:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] sync failed: {exc}")
        time.sleep(interval)


def main() -> None:
    parser = argparse.ArgumentParser(description="Sync an XLSX workbook into a Markdown table dump.")
    parser.add_argument("--source", type=Path, default=DEFAULT_XLSX, help="XLSX file to read.")
    parser.add_argument("--target", type=Path, default=DEFAULT_MD, help="Markdown file to write.")
    parser.add_argument("--watch", action="store_true", help="Keep running and sync whenever the XLSX changes.")
    parser.add_argument("--interval", type=float, default=2.0, help="Polling interval in seconds for --watch.")
    args = parser.parse_args()

    if args.watch:
        watch(args.source, args.target, args.interval)
    else:
        convert_xlsx_to_markdown(args.source, args.target)
        print(f"Synced {args.source} -> {args.target}")


if __name__ == "__main__":
    main()
