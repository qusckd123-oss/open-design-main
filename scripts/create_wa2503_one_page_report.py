from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import matplotlib.dates as mdates
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
from matplotlib.gridspec import GridSpec
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def make_rows(source_path: Path) -> tuple[list[dict], dict]:
    wb = load_workbook(source_path, data_only=True)
    ws = wb.active

    initial_qty = int(ws["I3"].value)
    sold_qty = int(ws["I4"].value)
    stock_qty = int(ws["I8"].value)
    sell_through = float(ws["I5"].value)
    review_date = ws["B16"].value or datetime(2026, 5, 19)
    first_sale_date = ws["B17"].value or datetime(2025, 6, 15)

    dates: list[datetime] = []
    source_weekly: list[float] = []
    for col in range(13, ws.max_column + 1):
        date_value = ws.cell(26, col).value
        weekly_value = ws.cell(27, col).value
        if isinstance(date_value, datetime) and isinstance(weekly_value, (int, float)):
            dates.append(date_value)
            source_weekly.append(float(weekly_value))

    reorder_1_date = datetime(2026, 5, 24)
    reorder_2_actual_date = datetime(2026, 6, 30)
    reorder_2_chart_week = datetime(2026, 7, 5)
    final_week = datetime(2026, 10, 4)
    reorder_1_qty = 3000
    reorder_2_qty = 7000
    total_supply = initial_qty + reorder_1_qty + reorder_2_qty

    seasonal_weights = {
        datetime(2026, 5, 24): 0.75,
        datetime(2026, 5, 31): 0.75,
        datetime(2026, 6, 7): 0.80,
        datetime(2026, 6, 14): 0.85,
        datetime(2026, 6, 21): 0.95,
        datetime(2026, 6, 28): 1.05,
        datetime(2026, 7, 5): 1.20,
        datetime(2026, 7, 12): 1.30,
        datetime(2026, 7, 19): 1.35,
        datetime(2026, 7, 26): 1.30,
        datetime(2026, 8, 2): 1.15,
        datetime(2026, 8, 9): 1.00,
        datetime(2026, 8, 16): 0.85,
        datetime(2026, 8, 23): 0.72,
        datetime(2026, 8, 30): 0.60,
        datetime(2026, 9, 6): 0.48,
        datetime(2026, 9, 13): 0.38,
        datetime(2026, 9, 20): 0.28,
        datetime(2026, 9, 27): 0.20,
        datetime(2026, 10, 4): 0.14,
    }
    future_weight_sum = sum(
        weight
        for date_value, weight in seasonal_weights.items()
        if reorder_1_date <= date_value <= final_week
    )
    forecast_unit = (total_supply - sold_qty) / future_weight_sum

    rows: list[dict] = []
    for date_value, source_value in zip(dates, source_weekly):
        if date_value < reorder_1_date:
            weekly = source_value
        else:
            weekly = seasonal_weights.get(date_value, 0) * forecast_unit
        rows.append({"date": date_value, "source_weekly": source_value, "weekly": weekly})

    historical_sum = sum(row["weekly"] for row in rows if row["date"] < reorder_1_date)
    historical_scale = sold_qty / historical_sum
    for row in rows:
        if row["date"] < reorder_1_date:
            row["weekly"] *= historical_scale

    cumulative_supply = initial_qty
    cumulative_sales = 0.0
    for row in rows:
        inbound = 0
        event = ""
        if row["date"] == reorder_1_date:
            inbound = reorder_1_qty
            event = "1차 리오더 3,000장 입고"
        if row["date"] == reorder_2_chart_week:
            inbound = reorder_2_qty
            event = "2차 리오더 7,000장 입고"

        cumulative_supply += inbound
        cumulative_sales += row["weekly"]
        if row["date"] == final_week:
            cumulative_sales = total_supply

        row["inbound"] = inbound
        row["event"] = event
        row["cumulative_supply"] = cumulative_supply
        row["cumulative_sales"] = cumulative_sales
        row["ending_stock"] = max(cumulative_supply - cumulative_sales, 0)
        row["sell_through_total"] = cumulative_sales / total_supply

    rows[-1]["ending_stock"] = 0
    rows[-1]["cumulative_sales"] = total_supply
    rows[-1]["sell_through_total"] = 1

    assumptions = {
        "initial_qty": initial_qty,
        "sold_qty": sold_qty,
        "stock_qty": stock_qty,
        "sell_through": sell_through,
        "review_date": review_date,
        "first_sale_date": first_sale_date,
        "reorder_1_date": reorder_1_date,
        "reorder_1_qty": reorder_1_qty,
        "reorder_2_actual_date": reorder_2_actual_date,
        "reorder_2_chart_week": reorder_2_chart_week,
        "reorder_2_qty": reorder_2_qty,
        "final_week": final_week,
        "total_supply": total_supply,
        "forecast_unit": forecast_unit,
        "historical_scale": historical_scale,
    }
    return rows, assumptions


def add_card(ax, x: float, y: float, w: float, h: float, title: str, value: str, note: str = "") -> None:
    ax.add_patch(
        plt.Rectangle(
            (x, y),
            w,
            h,
            transform=ax.transAxes,
            facecolor="#FFFFFF",
            edgecolor="#D8D0C0",
            linewidth=1.0,
        )
    )
    ax.text(x + 0.025, y + h - 0.085, title, transform=ax.transAxes, fontsize=9.5, color="#5E5A52")
    ax.text(x + 0.025, y + 0.145, value, transform=ax.transAxes, fontsize=18, color="#111111", weight="bold")
    if note:
        ax.text(x + 0.025, y + 0.045, note, transform=ax.transAxes, fontsize=9.0, color="#7A7468")


def create_png(rows: list[dict], assumptions: dict, output_png: Path) -> None:
    font_path = Path("C:/Windows/Fonts/malgun.ttf")
    if font_path.exists():
        fm.fontManager.addfont(str(font_path))
        plt.rcParams["font.family"] = "Malgun Gothic"
    plt.rcParams["axes.unicode_minus"] = False

    plot_rows = [row for row in rows if row["date"] >= datetime(2026, 3, 1)]
    dates = [row["date"] for row in plot_rows]
    stock = [row["ending_stock"] for row in plot_rows]
    sales = [row["cumulative_sales"] for row in plot_rows]
    supply = [row["cumulative_supply"] for row in plot_rows]
    weekly = [row["weekly"] for row in plot_rows]
    sell_through_pct = [row["sell_through_total"] * 100 for row in plot_rows]

    fig = plt.figure(figsize=(16, 9), dpi=180, facecolor="#F6F2EA")
    grid = GridSpec(8, 12, figure=fig, left=0.045, right=0.94, top=0.94, bottom=0.10, hspace=0.55)
    header_ax = fig.add_subplot(grid[0:2, :])
    chart_ax = fig.add_subplot(grid[3:8, :])
    header_ax.axis("off")

    header_ax.text(
        0.0,
        0.86,
        "WA2503STE1 TC 2차 리오더 진행 근거",
        fontsize=25,
        weight="bold",
        color="#111111",
        transform=header_ax.transAxes,
    )
    header_ax.text(
        0.0,
        0.67,
        "1차 리오더 3,000장 입고 후에도 6월 말 재고 런웨이가 짧아져, 2차 리오더 7,000장 진행 시 10/4주차 전체 소진 흐름",
        fontsize=12.5,
        color="#3C3933",
        transform=header_ax.transAxes,
    )
    header_ax.text(
        0.82,
        0.86,
        "대만 제외",
        fontsize=14,
        weight="bold",
        color="#FFFFFF",
        ha="center",
        va="center",
        bbox={"boxstyle": "round,pad=0.45", "fc": "#1F4E5F", "ec": "#1F4E5F"},
        transform=header_ax.transAxes,
    )

    add_card(header_ax, 0.00, 0.04, 0.18, 0.40, "현재 누적 판매", f"{assumptions['sold_qty']:,}장", "검토일 2026-05-19")
    add_card(header_ax, 0.20, 0.04, 0.18, 0.40, "현재 재고", f"{assumptions['stock_qty']:,}장", f"판매율 {assumptions['sell_through']:.1%}")
    add_card(header_ax, 0.40, 0.04, 0.18, 0.40, "1차 리오더", "3,000장", "2026-05-24 입고")
    add_card(header_ax, 0.60, 0.04, 0.18, 0.40, "2차 리오더 요청", "7,000장", "2026-06-30 입고")
    add_card(header_ax, 0.80, 0.04, 0.18, 0.40, "총 운영 공급", f"{assumptions['total_supply']:,}장", "10/4주차 100% 소진")

    chart_ax.set_facecolor("#F6F2EA")
    chart_ax.plot(dates, stock, color="#C43D36", linewidth=4.0, marker="o", markersize=4.5, label="기말 재고 런웨이")
    chart_ax.plot(dates, sales, color="#252525", linewidth=3.0, label="누적 판매량")
    chart_ax.plot(dates, supply, color="#2F5D7C", linewidth=3.0, linestyle="--", label="누적 공급량")
    chart_ax.plot(dates, weekly, color="#88ABA5", linewidth=2.1, alpha=0.75, label="주차별 판매/예측")
    percent_ax = chart_ax.twinx()
    percent_ax.plot(
        dates,
        sell_through_pct,
        color="#8A5FBF",
        linewidth=2.8,
        linestyle="-.",
        label="누적 소진율(%)",
    )

    event_styles = [
        (assumptions["reorder_1_date"], "5/24\n1차 +3,000장", 0.70),
        (assumptions["reorder_2_chart_week"], "6/30\n2차 +7,000장", 0.90),
        (assumptions["final_week"], "10/4주차\n전체 소진", 0.18),
    ]
    for date_value, label, y_ratio in event_styles:
        chart_ax.axvline(date_value, color="#8B5E34", linestyle=":", linewidth=1.6)
        is_final_event = date_value == assumptions["final_week"]
        chart_ax.annotate(
            label,
            xy=(date_value, assumptions["total_supply"] * y_ratio),
            xytext=(-8 if is_final_event else 8, 0),
            textcoords="offset points",
            fontsize=11.5,
            color="#3A2A1A",
            va="center",
            ha="right" if is_final_event else "left",
            bbox={"boxstyle": "round,pad=0.42", "fc": "#FFF8E5", "ec": "#C6A15B", "lw": 1.0},
        )

    chart_ax.fill_between(dates, stock, color="#C43D36", alpha=0.07)
    chart_ax.set_title("재고 런웨이 기준 리오더 필요 수량", loc="left", fontsize=17, weight="bold", pad=14)
    chart_ax.set_ylabel("수량", fontsize=11)
    percent_ax.set_ylabel("누적 소진율", fontsize=11)
    chart_ax.set_ylim(0, assumptions["total_supply"] * 1.12)
    percent_ax.set_ylim(0, 105)
    chart_ax.yaxis.set_major_formatter(lambda value, _pos: f"{int(value):,}")
    percent_ax.yaxis.set_major_formatter(lambda value, _pos: f"{int(value):.0f}%")
    chart_ax.xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=mdates.SU, interval=2))
    chart_ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))
    chart_ax.grid(axis="y", color="#D7CEBE", linewidth=0.9)
    chart_ax.spines["top"].set_visible(False)
    chart_ax.spines["right"].set_visible(False)
    percent_ax.spines["top"].set_visible(False)
    chart_lines, chart_labels = chart_ax.get_legend_handles_labels()
    percent_lines, percent_labels = percent_ax.get_legend_handles_labels()
    chart_ax.legend(
        chart_lines + percent_lines,
        chart_labels + percent_labels,
        loc="upper left",
        ncol=5,
        frameon=False,
        bbox_to_anchor=(0.0, -0.13),
        fontsize=10.5,
    )
    chart_ax.text(
        0.0,
        -0.22,
        "데이터 기준: 초도/판매/재고는 대만 제외 수량. 주차별 판매 예측은 5/17주차까지 실제 판매 데이터, 이후는 WA2502STE1 판매 흐름을 기반으로 반팔 시즌성을 일부 조정한 추정치.",
        transform=chart_ax.transAxes,
        fontsize=10,
        color="#5E5A52",
    )
    fig.savefig(output_png, bbox_inches="tight")
    plt.close(fig)


def create_xlsx(rows: list[dict], assumptions: dict, output_png: Path, output_xlsx: Path, source_path: Path) -> None:
    wb = Workbook()
    report_ws = wb.active
    report_ws.title = "보고용 1페이지"
    report_ws.sheet_view.showGridLines = False
    report_ws.page_setup.orientation = "landscape"
    report_ws.page_setup.fitToWidth = 1
    report_ws.page_setup.fitToHeight = 1
    report_ws.sheet_properties.pageSetUpPr.fitToPage = True
    report_ws.page_margins.left = 0.2
    report_ws.page_margins.right = 0.2
    report_ws.page_margins.top = 0.2
    report_ws.page_margins.bottom = 0.2

    for col in range(1, 14):
        report_ws.column_dimensions[get_column_letter(col)].width = 13
    for row in range(1, 35):
        report_ws.row_dimensions[row].height = 18

    image = XLImage(str(output_png))
    image.width = 1120
    image.height = 630
    report_ws.add_image(image, "A1")
    report_ws.print_area = "A1:M34"

    data_ws = wb.create_sheet("산출 데이터")
    data_ws.sheet_view.showGridLines = False
    headers = ["주차", "원천 주차 판매", "보정 주차 판매", "입고", "누적 공급", "누적 판매", "기말 재고", "누적 판매율", "이벤트"]
    for col, header in enumerate(headers, 1):
        cell = data_ws.cell(1, col)
        cell.value = header
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E5F")
        cell.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(rows, 2):
        values = [
            row["date"],
            row["source_weekly"],
            row["weekly"],
            row["inbound"],
            row["cumulative_supply"],
            row["cumulative_sales"],
            row["ending_stock"],
            row["sell_through_total"],
            row["event"],
        ]
        for col, value in enumerate(values, 1):
            data_ws.cell(row_idx, col).value = value
            data_ws.cell(row_idx, col).font = Font(name="Arial", size=9)
        data_ws.cell(row_idx, 1).number_format = "yyyy-mm-dd"
        for col in range(2, 8):
            data_ws.cell(row_idx, col).number_format = "#,##0"
        data_ws.cell(row_idx, 8).number_format = "0.0%"
    for col in range(1, 10):
        data_ws.column_dimensions[get_column_letter(col)].width = 18

    summary_ws = wb.create_sheet("요약")
    summary_ws["A1"] = "WA2503STE1 TC 리오더 분석 요약"
    summary_ws["A1"].font = Font(name="Arial", bold=True, size=16)
    summary = [
        ("원본 파일", source_path.name),
        ("기준", "대만 제외"),
        ("연간 운영", "25SS WA2502STE1 -> 25FW WA2503STE1 -> 26SS WA2602STE1"),
        ("1차 리오더 품번", "26SS WA2602STE1"),
        ("현재 누적 판매", assumptions["sold_qty"]),
        ("현재 재고", assumptions["stock_qty"]),
        ("초도 입고", assumptions["initial_qty"]),
        ("1차 리오더", assumptions["reorder_1_qty"]),
        ("1차 입고일", assumptions["reorder_1_date"]),
        ("2차 리오더", assumptions["reorder_2_qty"]),
        ("2차 입고일", assumptions["reorder_2_actual_date"]),
        ("전체 소진 예상", assumptions["final_week"]),
    ]
    for row_idx, (label, value) in enumerate(summary, 3):
        summary_ws.cell(row_idx, 1).value = label
        summary_ws.cell(row_idx, 2).value = value
        summary_ws.cell(row_idx, 1).font = Font(name="Arial", bold=True)
        summary_ws.cell(row_idx, 2).font = Font(name="Arial")
        if isinstance(value, datetime):
            summary_ws.cell(row_idx, 2).number_format = "yyyy-mm-dd"
        elif isinstance(value, (int, float)):
            summary_ws.cell(row_idx, 2).number_format = "#,##0"
    summary_ws.column_dimensions["A"].width = 22
    summary_ws.column_dimensions["B"].width = 70
    wb.save(output_xlsx)


def main() -> None:
    source_path = Path(os.environ["FILE"])
    output_png = source_path.parent / "WA2503STE1_TC_one_page_reorder_report.png"
    output_xlsx = source_path.parent / "WA2503STE1_TC_one_page_reorder_report.xlsx"

    rows, assumptions = make_rows(source_path)
    create_png(rows, assumptions, output_png)
    create_xlsx(rows, assumptions, output_png, output_xlsx, source_path)

    print(output_png)
    print(output_xlsx)


if __name__ == "__main__":
    main()
